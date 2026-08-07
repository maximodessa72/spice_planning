"""
Модуль экспорта результатов в Excel
Создаёт файл с форматированием по легенде EXCEL_LEGEND.md
"""

import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List
from datetime import datetime
from simulation import get_plan, get_bottleneck_recommendations  # get_plan — план позиции; get_bottleneck_recommendations — preview "Требует решения"
from data import N_MONTHS, get_current_year_month  # get_current_year_month() — свежая дата при каждом вызове


def get_month_label(mi: int) -> str:
    """Получить название месяца по индексу. Дата берётся напрямую при каждом вызове."""
    months = ['Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн', 'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']
    base_year, base_month = get_current_year_month()
    y = base_year + (base_month - 1 + mi) // 12
    m = (base_month - 1 + mi) % 12
    return f"{months[m]} {y}"


def format_number(num: int) -> str:
    """Форматировать число с разделителями"""
    return f"{num:,}".replace(",", " ")


def get_week_label(buf: float) -> str:
    """Получить метку недельного буфера"""
    if buf < 0.5:
        return "пред.мес"
    elif buf < 0.75:
        return "Нед.1"
    elif buf < 1.0:
        return "Нед.2"
    elif buf < 1.25:
        return "Нед.3"
    else:
        return "Нед.4"


def get_week_color(buf: float) -> str:
    """Получить цвет фона для метки недели"""
    if buf < 0.5:
        return "FF0000"  # Ярко красный (пред.мес)
    elif buf < 0.75:
        return "FFCCBC"  # Красный персиковый (Нед.1) - как буфер < 1.0
    elif buf < 1.0:
        return "FFECB3"  # Жёлтый (Нед.2) - как буфер 1.0-1.25
    elif buf < 1.25:
        return "D5F5E3"  # Зелёный (Нед.3) - как буфер >= 1.25
    else:
        return "1565C0"  # Синий (Нед.4)


def get_buffer_fill(buf: float) -> str:
    """Получить цвет заливки для буфера"""
    if buf < 1.0:
        return "FFCCBC"  # Персиковый
    elif buf < 1.25:
        return "FFECB3"  # Светло-жёлтый
    else:
        return "D5F5E3"  # Светло-зелёный


def create_excel(all_results: Dict[str, List[Dict]], groups: List[Dict], filename: str = "план_балансировка.xlsx"):
    """
    Создать Excel файл с результатами симуляции
    СТРУКТУРА КАК В СТАРОМ ФАЙЛЕ:
    - Колонка A: Група / Позиція
    - Колонка B: План (кг/міс)
    - Колонка C: Тип
    - Месяцы начинаются с колонки D (4)
    - Заголовки месяцев в строке 2
    - Подзаголовки в строке 4
    
    Args:
        all_results: результаты симуляции всех групп
        groups: список групп с данными
        filename: имя выходного файла
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "План + Балансування"
    
    # Стили
    thin = Side(style='thin', color='BBBBBB')
    thick = Side(style='medium', color='000000')  # Толстая чёрная граница между месяцами
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    bd_thick_right = Border(left=thin, right=thick, top=thin, bottom=thin)  # С толстой правой границей
    
    # Заголовок (строка 1)
    ws.merge_cells('A1:AY1')  # Покрываем все колонки (3 + 12*4 = 51 = AY)
    cell = ws['A1']
    today = datetime.now().strftime("%d %B %Y").replace("January", "січня").replace("February", "лютого").replace("March", "березня").replace("April", "квітня").replace("May", "травня").replace("June", "червня").replace("July", "липня").replace("August", "серпня").replace("September", "вересня").replace("October", "жовтня").replace("November", "листопада").replace("December", "грудня")
    # Автоматический диапазон дат
    start_month = get_month_label(0)  # Первый месяц
    end_month = get_month_label(N_MONTHS - 1)  # Последний месяц
    cell.value = f"План поставок + Балансування  |  Буфер моделі: г.4=1.2, решта=1.0  |  Мішок: 25 кг  |  {start_month} — {end_month}  |  {today}"
    cell.font = Font(name='Arial', bold=True, size=11, color='000000')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # СТРОКА 2 - ЗАГОЛОВКИ МЕСЯЦЕВ (начинаются с колонки 4)
    col = 4  # Месяцы начинаются с D (колонка 4)
    for mi in range(N_MONTHS):
        month_label = get_month_label(mi)
        
        # Объединяем 4 колонки для названия месяца
        start_col = get_column_letter(col)
        end_col = get_column_letter(col + 3)
        ws.merge_cells(f'{start_col}2:{end_col}2')
        
        cell = ws[f'{start_col}2']
        cell.value = month_label
        cell.font = Font(name='Arial', bold=True, size=10, color='000000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bd
        
        col += 4
    
    ws.row_dimensions[2].height = 20
    
    # СТРОКА 3 - ПУСТАЯ (или объединённые ячейки месяцев продолжаются)
    ws.row_dimensions[3].height = 5
    
    # СТРОКА 4 - ПОДЗАГОЛОВКИ КОЛОНОК
    subheader_row = 4
    
    # Колонка A - "Группа / Позиция"
    ws.merge_cells(f'A{subheader_row}:A{subheader_row}')
    cell = ws[f'A{subheader_row}']
    cell.value = "Группа / Позиция"
    cell.font = Font(name='Arial', bold=True, size=9, color='000000')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = bd
    ws.column_dimensions['A'].width = 38
    
    # Колонка B - "план закупок кг/мес"
    cell = ws[f'B{subheader_row}']
    cell.value = "план закупок\nкг/мес"
    cell.font = Font(name='Arial', bold=True, size=9, color='000000')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = bd
    ws.column_dimensions['B'].width = 9
    
    # Колонка C - "Тип"
    cell = ws[f'C{subheader_row}']
    cell.value = "Тип"
    cell.font = Font(name='Arial', bold=True, size=9, color='000000')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = bd
    ws.column_dimensions['C'].width = 7
    
    # Подзаголовки месяцев (начинаются с колонки 4)
    col = 4
    headers = ["Ост.нач\n(кг)", "Срок реал.\nнач (мес)", "Приход\n(кг)", "Заказ\n→ приход"]
    
    for mi in range(N_MONTHS):
        for i, h in enumerate(headers):
            c = ws.cell(row=subheader_row, column=col + i)
            c.value = h
            c.font = Font(name='Arial', bold=True, size=8, color='000000')
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # Последняя колонка месяца (Замовлення) - толстая правая граница
            if i == 3:
                c.border = bd_thick_right
            else:
                c.border = bd
            ws.column_dimensions[get_column_letter(col + i)].width = 13
        
        col += 4
    
    ws.row_dimensions[subheader_row].height = 30
    
    # Закрепление областей: колонки A, B, C и строки 1, 2, 3, 4
    # Замораживаем на ячейке D5 (первая незакреплённая ячейка)
    ws.freeze_panes = 'D5'
    
    # ДАННЫЕ ГРУПП
    data_row = 5
    
    for group in groups:
        group_name = group["name"]
        group_results = all_results[group_name]
        is_active = group.get("active", True)
        
        # "Требует решения" — preview-рекомендации (не влияют на расчёт, см. simulation.py)
        # Список эскалаций за весь горизонт -> словарь {месяц показа: рекомендация}
        bottleneck_recs_by_mi = {}
        if is_active:
            for rec in get_bottleneck_recommendations(group, group_results):
                bottleneck_recs_by_mi[rec["mi"]] = rec
        
        # СТРОКА ГРУППЫ
        # Колонка A - название группы
        cell = ws.cell(row=data_row, column=1)
        
        if is_active:
            # АКТИВНАЯ ГРУППА - синий фон, белый текст
            cell.value = group_name
            cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        else:
            # НЕАКТИВНАЯ ГРУППА - серый фон, серый текст, пометка
            cell.value = f"{group_name} (НЕАКТИВНА)"
            cell.font = Font(name='Arial', bold=True, size=9, color='999999')
            cell.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = bd
        
        # Колонка B - общий план группы
        cell = ws.cell(row=data_row, column=2)
        if is_active:
            total_plan = sum(it["plan"] for it in group["items"])
            cell.value = total_plan
            cell.number_format = '#,##0'
            cell.font = Font(name='Arial', bold=True, size=9, color='000000')
        else:
            cell.value = "—"
            cell.font = Font(name='Arial', bold=True, size=9, color='CCCCCC')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bd
        
        # Колонка C - "Гр." (группа)
        cell = ws.cell(row=data_row, column=3)
        cell.value = "Гр."
        if is_active:
            cell.font = Font(name='Arial', bold=True, size=8, color='000000')
            cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        else:
            cell.font = Font(name='Arial', bold=True, size=8, color='CCCCCC')
            cell.fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bd
        
        ws.row_dimensions[data_row].height = 34.0
        
        # Данные группы по месяцам (начинаются с колонки 4)
        col = 4
        for mi in range(N_MONTHS):
            r = group_results[mi]
            
            if is_active:
                # АКТИВНАЯ ГРУППА - обычное отображение
                # 1. Остаток на начало (с цветом буфера)
                total_balance = sum(r["bsi"].values())
                c = ws.cell(row=data_row, column=col)
                c.value = int(total_balance)
                c.number_format = '#,##0'
                c.font = Font(name='Arial', bold=False, size=9, color='363636')
                c.alignment = Alignment(horizontal='right', vertical='center')
                c.fill = PatternFill(start_color=get_buffer_fill(r["w_buf_before"]), 
                                    end_color=get_buffer_fill(r["w_buf_before"]), 
                                    fill_type='solid')
                c.border = bd
                
                # 2. Буфер (срок реализации) - ДО ПРИХОДА
                c = ws.cell(row=data_row, column=col + 1)
                buf_before = r["w_buf_before"]
                c.value = round(buf_before, 1)
                c.number_format = '0.0'
                c.font = Font(name='Arial', bold=False, size=9, color='000000')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.fill = PatternFill(start_color=get_buffer_fill(buf_before), 
                                    end_color=get_buffer_fill(buf_before), 
                                    fill_type='solid')
                c.border = bd
                
                # 3. Приход
                c = ws.cell(row=data_row, column=col + 2)
                if r["arrive"] > 0:
                    # ИСПОЛЬЗУЕМ week_label из симуляции (учитывает week_arrival)
                    week_label_full = r.get("wl", "")
                    
                    # Конвертируем метку из simulation.py в краткий формат
                    if "пред" in week_label_full.lower():
                        week_label = "пред.мес"
                    elif "Тиж. 1" in week_label_full:
                        week_label = "Нед.1"
                    elif "Тиж. 2" in week_label_full:
                        week_label = "Нед.2"
                    elif "Тиж. 3" in week_label_full:
                        week_label = "Нед.3"
                    elif "Тиж. 4" in week_label_full:
                        week_label = "Нед.4"
                    else:
                        # Fallback на старый метод если метка не распознана
                        week_label = get_week_label(r["w_buf_before"])
                    
                    # ФИКСИРОВАННЫЕ заказы (in_transit) - ВСЕГДА коричневые
                    if r["in_transit"]:
                        c.value = f"{int(r['arrive']):,}\n{week_label}".replace(",", " ")
                        c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
                        c.fill = PatternFill(start_color='5D4037', end_color='5D4037', fill_type='solid')
                        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    else:
                        # НОВЫЕ заказы - цвет из симуляции (учитывает week_arrival)
                        week_color = r.get("wc", "D5F5E3")  # Fallback на зелёный
                        
                        # Цвет текста: белый для тёмных фонов, чёрный для светлых
                        if week_color in ['FF0000', '1565C0']:
                            text_color = 'FFFFFF'
                        else:
                            text_color = '000000'
                        
                        c.value = f"{int(r['arrive']):,}\n{week_label}".replace(",", " ")
                        c.font = Font(name='Arial', bold=True, size=9, color=text_color)
                        c.fill = PatternFill(start_color=week_color, end_color=week_color, fill_type='solid')
                        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    c.value = "—"
                    c.font = Font(name='Arial', bold=False, size=9, color='CCCCCC')
                    c.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                    c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = bd
                
                # 4. Заказ
                c = ws.cell(row=data_row, column=col + 3)
                if r["containers"] > 0:
                    target_month = get_month_label(mi + math.ceil(group["cycle"] / 30))
                    c.value = f"{r['containers']} конт.\n→ {target_month}"
                    c.font = Font(name='Arial', bold=True, size=8, color='1565C0')
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    c.value = "—"
                    c.font = Font(name='Arial', bold=False, size=9, color='CCCCCC')
                    c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = bd_thick_right
            
            else:
                # НЕАКТИВНАЯ ГРУППА - все ячейки с прочерками, серый текст
                # 1. Остаток - прочерк
                c = ws.cell(row=data_row, column=col)
                c.value = "—"
                c.font = Font(name='Arial', bold=False, size=9, color='CCCCCC')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
                c.border = bd
                
                # 2. Буфер - прочерк
                c = ws.cell(row=data_row, column=col + 1)
                c.value = "—"
                c.font = Font(name='Arial', bold=False, size=9, color='CCCCCC')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
                c.border = bd
                
                # 3. Приход - прочерк
                c = ws.cell(row=data_row, column=col + 2)
                c.value = "—"
                c.font = Font(name='Arial', bold=False, size=9, color='CCCCCC')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
                c.border = bd
                
                # 4. Заказ - прочерк
                c = ws.cell(row=data_row, column=col + 3)
                c.value = "—"
                c.font = Font(name='Arial', bold=False, size=9, color='CCCCCC')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
                c.border = bd_thick_right
            
            col += 4
        
        data_row += 1
        
        # СТРОКИ ПОЗИЦИЙ
        # Для неактивных групп позиции не показываем
        if not is_active:
            continue
        
        for item_idx, item in enumerate(group["items"]):
            is_last_item = (item_idx == len(group["items"]) - 1)  # Последняя позиция?
            
            # Колонка A - название позиции с отступом (ГОЛУБОЙ ФОН, BOLD)
            cell = ws.cell(row=data_row, column=1)
            cell.value = f"   {item['name']}"  # Отступ 3 пробела
            cell.font = Font(name='Arial', bold=True, size=9, color='000000')  # BOLD
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.fill = PatternFill(start_color='EEF3FA', end_color='EEF3FA', fill_type='solid')  # Голубой фон
            # Толстая нижняя граница для последней позиции
            if is_last_item:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thick)
            else:
                cell.border = bd
            
            # Колонка B - план позиции (ГОЛУБОЙ ФОН)
            cell = ws.cell(row=data_row, column=2)
            # Если план=0 и есть plan_override - это спецпозиция
            if item["plan"] == 0 and item.get("plan_override"):
                cell.value = "★ спец"
            else:
                cell.value = item["plan"]
                cell.number_format = '#,##0'
            cell.font = Font(name='Arial', bold=False, size=9, color='000000')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='EEF3FA', end_color='EEF3FA', fill_type='solid')
            if is_last_item:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thick)
            else:
                cell.border = bd
            
            # Колонка C - тип (ГОЛУБОЙ ФОН)
            cell = ws.cell(row=data_row, column=3)
            # Определяем тип позиции
            if item["plan"] == 0 and item.get("plan_override"):
                item_type = "спец"
            elif item.get("seasonal"):
                item_type = "сез."
            else:
                item_type = "осн."
            cell.value = item_type
            cell.font = Font(name='Arial', bold=False, size=9, color='000000')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='EEF3FA', end_color='EEF3FA', fill_type='solid')
            if is_last_item:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thick)
            else:
                cell.border = bd
            
            ws.row_dimensions[data_row].height = 20.0
            
            # Данные позиции по месяцам (начинаются с колонки 4)
            col = 4
            for mi in range(N_MONTHS):
                r = group_results[mi]
                
                # "Требует решения" — эта позиция в этом месяце попала в preview-рекомендацию?
                rec_this_month = bottleneck_recs_by_mi.get(mi)
                is_rec = (
                    rec_this_month is not None
                    and rec_this_month["order_kg"].get(item["name"], 0) > 0
                )
                
                # 1. Остаток позиции на начало (с цветом буфера)
                c = ws.cell(row=data_row, column=col)
                c.value = int(r["bsi"][item["name"]])
                c.number_format = '#,##0'
                c.font = Font(name='Arial', bold=False, size=9, color='363636')
                c.alignment = Alignment(horizontal='right', vertical='center')
                # Применяем цвет буфера ДО прихода
                buf_val_for_color = r["icb"][item["name"]]
                c.fill = PatternFill(start_color=get_buffer_fill(buf_val_for_color), 
                                    end_color=get_buffer_fill(buf_val_for_color), 
                                    fill_type='solid')
                if is_last_item:
                    c.border = Border(left=thin, right=thin, top=thin, bottom=thick)
                else:
                    c.border = bd
                
                # 2. Буфер позиции ДО прихода
                c = ws.cell(row=data_row, column=col + 1)
                plan_val = get_plan(item, mi)
                if plan_val == 0:
                    # Если план=0, показываем прочерк
                    c.value = "—"
                    c.font = Font(name='Arial', bold=False, size=9, color='CCCCCC')
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                else:
                    buf_val = r["icb"][item["name"]]  # Буфер ДО прихода (НЕ округлённый)
                    c.value = round(buf_val, 1)  # Показываем округлённое
                    c.number_format = '0.0'
                    c.font = Font(name='Arial', bold=True, size=9, color='000000')
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    # Применяем цвет по НЕ округлённому значению
                    c.fill = PatternFill(start_color=get_buffer_fill(buf_val), 
                                        end_color=get_buffer_fill(buf_val), 
                                        fill_type='solid')
                if is_last_item:
                    c.border = Border(left=thin, right=thin, top=thin, bottom=thick)
                else:
                    c.border = bd
                
                # 3. Приход позиции
                c = ws.cell(row=data_row, column=col + 2)
                if r["ia"][item["name"]] > 0:
                    # Если это фиксированный заказ - КОРИЧНЕВЫЙ фон
                    if r["in_transit"]:
                        c.value = int(r["ia"][item["name"]])
                        c.number_format = '#,##0'
                        c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
                        c.fill = PatternFill(start_color='5D4037', end_color='5D4037', fill_type='solid')
                        c.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        # Новый заказ - ЗЕЛЁНЫЙ фон
                        c.value = int(r["ia"][item["name"]])
                        c.number_format = '#,##0'
                        c.font = Font(name='Arial', bold=False, size=9, color='000000')
                        c.fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
                        c.alignment = Alignment(horizontal='center', vertical='center')
                elif is_rec:
                    # "Требует решения" - ФИОЛЕТОВЫЙ фон, это рекомендация, не заказ
                    rec_kg = rec_this_month["order_kg"][item["name"]]
                    c.value = f"{int(rec_kg):,}\nтреб.реш.".replace(",", " ")
                    c.number_format = '@'
                    c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
                    c.fill = PatternFill(start_color='7B1FA2', end_color='7B1FA2', fill_type='solid')
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    c.value = "—"
                    c.font = Font(name='Arial', bold=False, size=9, color='CCCCCC')
                    # Нет прихода - голубой фон
                    c.fill = PatternFill(start_color='EEF3FA', end_color='EEF3FA', fill_type='solid')
                    c.alignment = Alignment(horizontal='center', vertical='center')
                if is_last_item:
                    c.border = Border(left=thin, right=thin, top=thin, bottom=thick)
                else:
                    c.border = bd
                
                # 4. Буфер ПОСЛЕ прихода (в колонке Замовлення)
                c = ws.cell(row=data_row, column=col + 3)
                plan_val = get_plan(item, mi)
                if plan_val == 0:
                    # Если план=0, показываем прочерк
                    c.value = "—"
                    c.font = Font(name='Arial', bold=False, size=9, color='CCCCCC')
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                elif is_rec:
                    # "Требует решения" - буфер ПОСЛЕ рекомендации, фиолетовый
                    buf_after_rec = rec_this_month["buf_after"][item["name"]]
                    c.value = round(buf_after_rec, 1)
                    c.number_format = '0.0'
                    c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.fill = PatternFill(start_color='7B1FA2', end_color='7B1FA2', fill_type='solid')
                else:
                    buf_val_after = r["ica"][item["name"]]  # Буфер ПОСЛЕ (НЕ округлённый)
                    c.value = round(buf_val_after, 1)  # Показываем округлённое
                    c.number_format = '0.0'
                    c.font = Font(name='Arial', bold=False, size=9, color='000000')
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    # Применяем цвет по НЕ округлённому значению
                    c.fill = PatternFill(start_color=get_buffer_fill(buf_val_after), 
                                        end_color=get_buffer_fill(buf_val_after), 
                                        fill_type='solid')
                # Толстая правая граница для разделения месяцев + толстая нижняя для последней позиции
                if is_last_item:
                    c.border = Border(left=thin, right=thick, top=thin, bottom=thick)
                else:
                    c.border = bd_thick_right
                
                col += 4
            
            data_row += 1
    
    # ========================================================================
    # ЗАКЛАДКА: КАЛЕНДАРЬ ЗАКАЗОВ
    # ========================================================================
    
    ws_calendar = wb.create_sheet("Календарь заказов")
    
    # Заголовок
    ws_calendar.merge_cells('A1:N1')
    cell = ws_calendar['A1']
    cell.value = "Календарь заказов на 12 месяцев"
    cell.font = Font(name='Arial', bold=True, size=12, color='000000')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_calendar.row_dimensions[1].height = 25
    
    # Заголовки колонок
    headers_row = 3
    ws_calendar['A3'] = "Группа"
    ws_calendar['A3'].font = Font(name='Arial', bold=True, size=10)
    ws_calendar['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws_calendar.column_dimensions['A'].width = 35
    
    # Месяцы (колонки B-M)
    for mi in range(N_MONTHS):
        col_letter = get_column_letter(2 + mi)  # B=2, C=3, etc.
        month_label = get_month_label(mi)
        cell = ws_calendar[f'{col_letter}3']
        cell.value = month_label
        cell.font = Font(name='Arial', bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_calendar.column_dimensions[col_letter].width = 12
    
    ws_calendar.row_dimensions[3].height = 20
    
    # Данные по группам
    data_row = 4
    
    for group in groups:
        group_name = group["name"]
        group_results = all_results[group_name]
        
        # Колонка A - название группы
        cell = ws_calendar.cell(row=data_row, column=1)
        cell.value = group_name
        cell.font = Font(name='Arial', bold=False, size=9)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Колонки B-M - месяцы прибытия
        for mi in range(N_MONTHS):
            r = group_results[mi]
            cell = ws_calendar.cell(row=data_row, column=2 + mi)
            
            if r["containers"] > 0:
                # Есть заказ - показываем месяц прибытия
                arrival_mi = mi + math.ceil(group["cycle"] / 30)
                arrival_month = get_month_label(arrival_mi)
                
                # Определяем цвет
                if r["in_transit"]:
                    # Фиксированный заказ - КОРИЧНЕВЫЙ
                    bg_color = '5D4037'
                    text_color = 'FFFFFF'
                else:
                    # Новый заказ - цвет по метке недели
                    week_color = get_week_color(r["w_buf_before"])
                    bg_color = week_color
                    
                    # Цвет текста: белый для тёмных фонов, чёрный для светлых
                    if week_color in ['FF0000', '1565C0']:
                        text_color = 'FFFFFF'
                    else:
                        text_color = '000000'
                
                cell.value = arrival_month
                cell.font = Font(name='Arial', bold=True, size=9, color=text_color)
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # Нет заказа
                cell.value = ""
            
            cell.border = bd
        
        ws_calendar.row_dimensions[data_row].height = 18
        data_row += 1
    
    # ========================================================================
    # ЛЕГЕНДА
    # ========================================================================
    
    # Пропускаем 2 строки
    legend_row = data_row + 2
    
    # Заголовок легенды
    ws_calendar.merge_cells(f'A{legend_row}:D{legend_row}')
    cell = ws_calendar[f'A{legend_row}']
    cell.value = "ЛЕГЕНДА:"
    cell.font = Font(name='Arial', bold=True, size=10, color='000000')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    
    legend_row += 1
    
    # Фиксированные заказы
    cell = ws_calendar.cell(row=legend_row, column=1)
    cell.value = "Коричневый"
    cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
    cell.fill = PatternFill(start_color='5D4037', end_color='5D4037', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    cell = ws_calendar.cell(row=legend_row, column=2)
    cell.value = "Фиксированный заказ (уже в пути)"
    cell.font = Font(name='Arial', bold=False, size=9)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    
    ws_calendar.row_dimensions[legend_row].height = 18
    legend_row += 1
    
    # Новые заказы
    ws_calendar.merge_cells(f'A{legend_row}:D{legend_row}')
    cell = ws_calendar[f'A{legend_row}']
    cell.value = "Новые заказы (по срочности):"
    cell.font = Font(name='Arial', bold=True, size=9, color='000000')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    
    legend_row += 1
    
    # Таблица меток
    legend_items = [
        ('пред.мес', 'FF0000', 'FFFFFF', 'Критично! Буфер < 0.5 мес'),
        ('Нед.1', 'FFCCBC', '000000', 'Срочно. Буфер 0.5-0.75 мес'),
        ('Нед.2', 'FFECB3', '000000', 'Внимание. Буфер 0.75-1.0 мес'),
        ('Нед.3', 'D5F5E3', '000000', 'Нормально. Буфер 1.0-1.25 мес'),
        ('Нед.4', '1565C0', 'FFFFFF', 'Хорошо. Буфер >= 1.25 мес')
    ]
    
    for label, bg_color, text_color, description in legend_items:
        # Колонка A - метка с цветом
        cell = ws_calendar.cell(row=legend_row, column=1)
        cell.value = label
        cell.font = Font(name='Arial', bold=True, size=9, color=text_color)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Колонка B - описание
        cell = ws_calendar.cell(row=legend_row, column=2)
        cell.value = description
        cell.font = Font(name='Arial', bold=False, size=9)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        ws_calendar.row_dimensions[legend_row].height = 18
        legend_row += 1
    
    # "Требует решения" (preview-рекомендация, не заказ)
    cell = ws_calendar.cell(row=legend_row, column=1)
    cell.value = "Фиолетовый"
    cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
    cell.fill = PatternFill(start_color='7B1FA2', end_color='7B1FA2', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    cell = ws_calendar.cell(row=legend_row, column=2)
    cell.value = "Требует решения — узкое место по отдельной позиции. Это рекомендация, не заказ."
    cell.font = Font(name='Arial', bold=False, size=9)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    
    ws_calendar.row_dimensions[legend_row].height = 18
    legend_row += 1
    
    # ========================================================================
    # ЗАКЛАДКА: КАЛЕНДАРЬ ПОСТАВОК
    # ========================================================================
    
    ws_arrivals = wb.create_sheet("Календарь поставок")
    
    # Заголовок
    ws_arrivals.merge_cells('A1:N1')
    cell = ws_arrivals['A1']
    cell.value = "Календарь поставок (приходы контейнеров)"
    cell.font = Font(name='Arial', bold=True, size=12, color='000000')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_arrivals.row_dimensions[1].height = 25
    
    # Заголовки колонок
    ws_arrivals['A3'] = "Группа"
    ws_arrivals['A3'].font = Font(name='Arial', bold=True, size=10)
    ws_arrivals['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws_arrivals.column_dimensions['A'].width = 35
    
    # Месяцы (колонки B-M)
    for mi in range(N_MONTHS):
        col_letter = get_column_letter(2 + mi)
        month_label = get_month_label(mi)
        cell = ws_arrivals[f'{col_letter}3']
        cell.value = month_label
        cell.font = Font(name='Arial', bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_arrivals.column_dimensions[col_letter].width = 10
    
    # Колонка ВСЕГО
    col_letter = get_column_letter(14)  # N
    cell = ws_arrivals[f'{col_letter}3']
    cell.value = "ВСЕГО"
    cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_arrivals.column_dimensions[col_letter].width = 10
    
    ws_arrivals.row_dimensions[3].height = 20
    
    # Данные по группам
    data_row = 4
    month_totals = {mi: 0 for mi in range(N_MONTHS)}
    grand_total_arrivals = 0
    
    for group in groups:
        group_name = group["name"]
        group_results = all_results[group_name]
        
        # Колонка A - название группы
        cell = ws_arrivals.cell(row=data_row, column=1)
        cell.value = group_name
        cell.font = Font(name='Arial', bold=False, size=9)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = bd
        
        row_total = 0
        
        # Колонки B-M - количество контейнеров
        for mi in range(N_MONTHS):
            r = group_results[mi]
            cell = ws_arrivals.cell(row=data_row, column=2 + mi)
            
            if r["arrive"] > 0:
                # Используем unit_container если есть, иначе container
                container_size = group.get("unit_container", group["container"])
                num_containers = round(r["arrive"] / container_size)
                cell.value = num_containers
                
                # Цвет: коричневый если заказан, зелёный если новый
                if r["in_transit"]:
                    # Заказан - коричневый
                    cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
                    cell.fill = PatternFill(start_color='5D4037', end_color='5D4037', fill_type='solid')
                else:
                    # Новый - зелёный
                    cell.font = Font(name='Arial', bold=True, size=9, color='000000')
                    cell.fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
                
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                row_total += num_containers
                month_totals[mi] += num_containers
                grand_total_arrivals += num_containers
            else:
                cell.value = ""
            
            cell.border = bd
        
        # Колонка ВСЕГО для группы
        cell = ws_arrivals.cell(row=data_row, column=14)
        if row_total > 0:
            cell.value = row_total
            cell.font = Font(name='Arial', bold=True, size=9)
            cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        else:
            cell.value = ""
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bd
        
        ws_arrivals.row_dimensions[data_row].height = 18
        data_row += 1
    
    # Строка ИТОГО
    cell = ws_arrivals.cell(row=data_row, column=1)
    cell.value = "ИТОГО"
    cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = bd
    
    for mi in range(N_MONTHS):
        cell = ws_arrivals.cell(row=data_row, column=2 + mi)
        if month_totals[mi] > 0:
            cell.value = month_totals[mi]
            cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        else:
            cell.value = ""
            cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bd
    
    # ВСЕГО-ВСЕГО
    cell = ws_arrivals.cell(row=data_row, column=14)
    cell.value = grand_total_arrivals
    cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = bd
    
    ws_arrivals.row_dimensions[data_row].height = 22
    
    # Сохранение
    wb.save(filename)
    return filename
