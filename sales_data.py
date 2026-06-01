"""
Данные для модуля планирования продаж
Использует общие данные из data.py (GROUPS и т.д.)
"""

from datetime import datetime
import openpyxl
from typing import Dict

# Импортируем общие данные
from data import GROUPS

# КОНСТАНТЫ ДЛЯ ПРОДАЖ (одинаковые с закупками!)
SALES_N_MONTHS = 12  # Планирование на 12 месяцев вперёд

# АВТООПРЕДЕЛЕНИЕ текущего месяца (как в закупках)
_current_date = datetime.now()
SALES_BASE_YEAR = _current_date.year
SALES_BASE_MONTH = _current_date.month  # Текущий месяц


# ========== ФУНКЦИИ ИМПОРТА ==========

def import_prices_from_excel(filepath: str) -> Dict:
    """
    Импорт актуальных цен из Excel
    
    Формат файла:
    - Строка 2: Заголовки (Группа / Позиция, Актуальная цена, грн)
    - Строка 4+: Группы и позиции
    
    Returns:
        {group_idx: {item_idx: price}}
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    prices = {}
    current_group_idx = -1
    current_item_idx = -1
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
        name = row[0]
        price = row[1]
        
        if not name:
            continue
        
        # Это группа (без отступа)
        if not name.startswith('   '):
            # Находим индекс группы в GROUPS
            current_group_idx = -1
            for gi, g in enumerate(GROUPS):
                if g["name"] == name:
                    current_group_idx = gi
                    break
            current_item_idx = -1
        else:
            # Это позиция (с отступом)
            item_name = name.strip()
            
            if current_group_idx == -1:
                continue
            
            # Находим индекс позиции в группе
            group = GROUPS[current_group_idx]
            for ii, item in enumerate(group["items"]):
                if item["name"] == item_name:
                    current_item_idx = ii
                    break
            
            if current_item_idx == -1:
                continue
            
            # Сохраняем цену
            if price is not None and price > 0:
                if current_group_idx not in prices:
                    prices[current_group_idx] = {}
                prices[current_group_idx][current_item_idx] = float(price)
    
    return prices


def import_sales_plan_from_excel(filepath: str) -> Dict:
    """
    Импорт базового плана продаж из Excel
    
    Формат файла:
    - Строка 2: Заголовки (Группа / Позиция, План продаж базовый, кг)
    - Строка 4+: Группы и позиции
    
    Returns:
        {group_idx: {item_idx: plan_kg}}
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    plans = {}
    current_group_idx = -1
    current_item_idx = -1
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
        name = row[0]
        plan_kg = row[1]
        
        if not name:
            continue
        
        # Это группа (без отступа)
        if not name.startswith('   '):
            # Находим индекс группы в GROUPS
            current_group_idx = -1
            for gi, g in enumerate(GROUPS):
                if g["name"] == name:
                    current_group_idx = gi
                    break
            current_item_idx = -1
        else:
            # Это позиция (с отступом)
            item_name = name.strip()
            
            if current_group_idx == -1:
                continue
            
            # Находим индекс позиции в группе
            group = GROUPS[current_group_idx]
            for ii, item in enumerate(group["items"]):
                if item["name"] == item_name:
                    current_item_idx = ii
                    break
            
            if current_item_idx == -1:
                continue
            
            # Сохраняем план
            try:
                plan_kg = int(float(str(plan_kg).replace(',', '.').strip())) if plan_kg is not None else None
            except (ValueError, TypeError):
                plan_kg = None
            if plan_kg is not None and plan_kg > 0:
                if current_group_idx not in plans:
                    plans[current_group_idx] = {}
                plans[current_group_idx][current_item_idx] = int(plan_kg)
    
    return plans

# ========== АКТУАЛЬНЫЕ ЦЕНЫ ==========
# Структура: {group_index: {item_index: {month_index: price}}}
# Цены обновляются первого числа каждого месяца
SALES_PRICES = {}

# Пример структуры (будет заполняться из Excel):
# SALES_PRICES = {
#     0: {  # Группа 1. Єгипет (трави)
#         0: {0: 150.00, 1: 152.50},  # Базилік зелень (В) - цены по месяцам
#         1: {0: 145.00, 1: 147.00},  # Базилік зелень (С)
#     }
# }


# ========== ПЛАН ПРОДАЖ БАЗОВЫЙ ==========
# Структура: {group_index: {item_index: {month_index: plan_kg}}}
# План загружается первого числа каждого месяца
SALES_PLAN_BASE = {}

# Пример структуры (будет заполняться из Excel):
# SALES_PLAN_BASE = {
#     0: {  # Группа 1. Єгипет (трави)
#         0: {0: 5000, 1: 5200},  # Базилік зелень (В) - план продаж в кг по месяцам
#         1: {0: 1400, 1: 1500},  # Базилік зелень (С)
#     }
# }


# ========== ФАКТ ПРОДАЖ ==========
# Структура: {group_index: {item_index: {month_index: fact_kg}}}
# Факт загружается первого числа следующего месяца (в прошлый месяц -1)
SALES_FACT = {}

# Пример структуры (будет заполняться из Excel):
# SALES_FACT = {
#     0: {  # Группа 1. Єгипет (трави)
#         0: {0: 4850},  # Базилік зелень (В) - фактические продажи в кг
#         1: {0: 1520},  # Базилік зелень (С) - перевыполнение плана
#     }
# }


# ========== ДАТЫ ПРИХОДА ЗАКАЗОВ ==========
# Структура: {group_index: {month_index: {"date": "YYYY-MM-DD", "week": 1-4}}}
# Используется для корректировки плана продаж
ARRIVAL_DATES = {}

# Пример структуры:
# ARRIVAL_DATES = {
#     0: {  # Группа 1. Єгипет (трави)
#         0: {"date": "2026-04-08", "week": 2},  # Приход во 2-ю неделю апреля
#         1: {"date": "2026-05-23", "week": 4},  # Приход в 4-ю неделю мая
#     }
# }


# ========== КОЭФФИЦИЕНТЫ КОРРЕКТИРОВКИ ПЛАНА ==========
# В зависимости от недели прихода товара корректируем план продаж
# Если остаток < план и товар приходит в течение месяца
WEEK_ADJUSTMENT_COEFFICIENTS = {
    1: 0.90,  # 1-я неделя — план корректируется на 90% (больше времени на продажи)
    2: 0.70,  # 2-я неделя — план корректируется на 70%
    3: 0.50,  # 3-я неделя — план корректируется на 50%
    4: 0.30,  # 4-я неделя — план корректируется на 30% (мало времени на продажи)
}


def get_adjusted_sales_plan(group_idx: int, item_idx: int, month_idx: int, 
                            opening_balance: float, base_plan: float) -> float:
    """
    Получить скорректированный план продаж с учётом остатка и даты прихода
    
    Args:
        group_idx: индекс группы
        item_idx: индекс позиции в группе
        month_idx: индекс месяца
        opening_balance: остаток на начало месяца
        base_plan: базовый план продаж
    
    Returns:
        Скорректированный план продаж (кг)
    """
    # Если остаток >= плана — корректировка не нужна
    if opening_balance >= base_plan:
        return base_plan
    
    # Если есть фиксированная дата прихода в этом месяце
    if group_idx in ARRIVAL_DATES and month_idx in ARRIVAL_DATES[group_idx]:
        arrival_info = ARRIVAL_DATES[group_idx][month_idx]
        week = arrival_info.get("week", 4)
        
        # Получаем коэффициент корректировки
        adjustment = WEEK_ADJUSTMENT_COEFFICIENTS.get(week, 0.30)
        
        # Корректируем план: остаток + (недостающая часть * коэффициент)
        shortage = base_plan - opening_balance
        adjusted_plan = opening_balance + (shortage * adjustment)
        
        return adjusted_plan
    
    # Если даты прихода нет — возвращаем остаток (консервативный подход)
    return opening_balance


def get_sales_performance(group_idx: int, item_idx: int, month_idx: int) -> dict:
    """
    Получить показатели выполнения плана продаж
    
    Returns:
        {
            "plan": план продаж,
            "fact": факт продаж,
            "delta": отклонение (факт - план),
            "performance_pct": выполнение плана в %,
            "price": цена продажи
        }
    """
    plan = SALES_PLAN_BASE.get(group_idx, {}).get(item_idx, {}).get(month_idx, 0)
    fact = SALES_FACT.get(group_idx, {}).get(item_idx, {}).get(month_idx, 0)
    price = SALES_PRICES.get(group_idx, {}).get(item_idx, {}).get(month_idx, 0)
    
    delta = fact - plan
    performance_pct = (fact / plan * 100) if plan > 0 else 0
    
    return {
        "plan": plan,
        "fact": fact,
        "delta": delta,
        "performance_pct": performance_pct,
        "price": price
    }
