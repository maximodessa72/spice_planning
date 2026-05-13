"""
Экспорт плана продаж в Excel
Создаёт файл с планированием продаж на основе данных закупок
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from typing import Dict, List
from data import GROUPS
from sales_data import SALES_N_MONTHS, SALES_BASE_YEAR, SALES_BASE_MONTH
from simulation import get_plan  # Для получения сезонного плана


# Константы для расчёта рабочих дней
WORKING_DAYS_BY_WEEK = {
    "пред мес": 21,  # товар пришёл в прошлом месяце → доступен весь месяц
    "нед 1": 21,     # приход в 1-ю неделю → товар доступен весь месяц
    "нед 2": 16,     # приход во 2-ю неделю → с Пн 2-й недели
    "нед 3": 11,     # приход в 3-ю неделю → с Пн 3-й недели
    "нед 4": 5,      # приход в 4-ю неделю → с Пн 4-й недели
    None: 21,        # дата не зафиксирована → оптимистично, как нед 1
}


def get_month_label(mi: int) -> str:
    """Получить название месяца по индексу"""
    months = ['Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн', 'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']
    
    y = SALES_BASE_YEAR + (SALES_BASE_MONTH - 1 + mi) // 12
    m = (SALES_BASE_MONTH - 1 + mi) % 12
    
    return f"{months[m]} {y}"


def calculate_working_days_exact(arrival_date: datetime, month_end: datetime) -> int:
    """
    Точный расчёт рабочих дней (Пн-Пт) с даты прихода до конца месяца
    День прихода НЕ считается (выгрузка)
    
    Args:
        arrival_date: дата прихода контейнера
        month_end: последний день месяца
    
    Returns:
        Количество рабочих дней
    """
    working_days = 0
    # Начинаем со следующего дня после прихода
    current = arrival_date + timedelta(days=1)
    
    while current <= month_end:
        # 0=Пн, 1=Вт, 2=Ср, 3=Чт, 4=Пт, 5=Сб, 6=Вс
        if current.weekday() < 5:  # Пн-Пт
            working_days += 1
        current += timedelta(days=1)
    
    return working_days


def get_working_days_for_arrival(arrival_info: str, month_index: int, 
                                 current_month_index: int, 
                                 fixed_dates: Dict = None) -> int:
    """
    Получить количество рабочих дней с момента прихода до конца месяца
    
    Args:
        arrival_info: строка прихода ("1 конт.\\n→ Авг 2026" или "нед 2")
        month_index: индекс месяца (0-11)
        current_month_index: индекс текущего месяца
        fixed_dates: словарь зафиксированных дат {group_idx: {month_idx: datetime}}
    
    Returns:
        Количество рабочих дней
    """
    fixed_dates = fixed_dates or {}
    
    # Если это текущий месяц и есть зафиксированная дата
    if month_index == current_month_index and fixed_dates:
        # TODO: здесь будет логика получения зафиксированной даты
        # Пока возвращаем None (не зафиксирована)
        pass
    
    # Парсим информацию о неделе из arrival_info
    if "нед" in arrival_info.lower():
        # Извлекаем номер недели
        week_num = None
        for i in range(1, 5):
            if f"нед {i}" in arrival_info.lower() or f"нед. {i}" in arrival_info.lower():
                week_num = i
                break
        
        if week_num:
            return WORKING_DAYS_BY_WEEK[f"нед {week_num}"]
    
    # Если "пред мес"
    if "пред" in arrival_info.lower():
        return WORKING_DAYS_BY_WEEK["пред мес"]
    
    # По умолчанию (не зафиксирована дата) - оптимистично
    return WORKING_DAYS_BY_WEEK[None]


def calculate_sales_plan(opening_balance: float, arrival_kg: float, base_plan: float, 
                         working_days: int) -> int:
    """
    Рассчитать план продаж с учётом остатка, прихода и рабочих дней
    
    Формула: min(
        Остаток + Приход,
        Остаток + (Рабочие_дни × План_на_день),
        План_базовый
    )
    где План_на_день = План_базовый / 21
    
    Args:
        opening_balance: остаток на начало месяца (кг)
        arrival_kg: приход товара (кг)
        base_plan: базовый план продаж (кг/мес)
        working_days: количество рабочих дней после прихода
    
    Returns:
        План продаж (кг, округлённый до целых)
    """
    if base_plan == 0:
        return 0
    
    # Максимум что можем продать = остаток + приход
    max_available = opening_balance + arrival_kg
    
    # План на один день
    plan_per_day = base_plan / 21
    
    # Рассчитываем план с учётом рабочих дней
    calculated_plan = opening_balance + (working_days * plan_per_day)
    
    # План не может быть больше: базового плана, доступного товара, расчётного плана
    final_plan = min(calculated_plan, base_plan, max_available)
    
    # Округляем до целых
    return round(final_plan)


def calculate_working_days_for_month(mi: int, group_name: str, week_label: str, 
                                     arrival_kg: float, arrival_fixed_dates: Dict,
                                     group: Dict = None) -> int:
    """
    Рассчитать рабочие дни для месяца с учётом зафиксированных дат
    
    Args:
        mi: индекс месяца (для продаж: 0=прошлый, 1=текущий)
        group_name: название группы
        week_label: метка недели из симуляции
        arrival_kg: количество прихода
        arrival_fixed_dates: зафиксированные даты {group_name: datetime}
        group: данные группы (для week_arrival)
    
    Returns:
        Количество рабочих дней
    """
    if arrival_kg <= 0:
        return 0
    
    # Текущий месяц теперь mi=0 (май)
    is_current_month = (mi == 0)
    
    # Если текущий месяц и есть зафиксированная дата — считаем точно
    if is_current_month and arrival_fixed_dates and group_name in arrival_fixed_dates:
        from datetime import datetime
        from sales_data import SALES_BASE_YEAR, SALES_BASE_MONTH
        import calendar
        
        fixed_date = arrival_fixed_dates[group_name]
        
        # Определяем последний день месяца
        current_year = SALES_BASE_YEAR + (SALES_BASE_MONTH - 1 + mi) // 12
        current_month = (SALES_BASE_MONTH - 1 + mi) % 12 + 1
        last_day = calendar.monthrange(current_year, current_month)[1]
        month_end = datetime(current_year, current_month, last_day)
        
        # Рассчитываем точные рабочие дни
        return calculate_working_days_exact(fixed_date, month_end)
    else:
        # ПРИОРИТЕТ 1: Проверяем зафиксированный номер недели в week_arrival
        if group and "week_arrival" in group:
            week_num = group["week_arrival"].get(mi)
            if week_num:
                # Используем номер недели из week_arrival
                week_key = f"нед {week_num}"
                return WORKING_DAYS_BY_WEEK.get(week_key, WORKING_DAYS_BY_WEEK[None])
        
        # ПРИОРИТЕТ 2: Для будущих месяцев используем метку из симуляции
        if "Тиж. 1" in week_label or "пред" in week_label.lower():
            return WORKING_DAYS_BY_WEEK["нед 1"]
        elif "Тиж. 2" in week_label:
            return WORKING_DAYS_BY_WEEK["нед 2"]
        elif "Тиж. 3" in week_label:
            return WORKING_DAYS_BY_WEEK["нед 3"]
        elif "Тиж. 4" in week_label:
            return WORKING_DAYS_BY_WEEK["нед 4"]
        else:
            return WORKING_DAYS_BY_WEEK[None]


def create_sales_excel(procurement_results: Dict[str, List[Dict]], 
                       groups: List[Dict],
                       sales_prices: Dict = None,
                       sales_plan_base: Dict = None,
                       sales_fact: Dict = None,
                       arrival_fixed_dates: Dict = None,
                       filename: str = "план_продаж.xlsx",
                       group_limit: int = None):
    """
    Создать Excel файл с планом продаж
    
    Args:
        procurement_results: результаты симуляции закупок
        groups: список групп товаров
        sales_prices: актуальные цены {group_idx: {item_idx: {month_idx: price}}}
        sales_plan_base: базовый план продаж {group_idx: {item_idx: plan_kg}}
        sales_fact: факт продаж {group_idx: {item_idx: {month_idx: fact_kg}}}
        arrival_fixed_dates: зафиксированные даты прихода
        filename: имя выходного файла
        group_limit: ограничение количества групп (для теста)
    """
    sales_prices = sales_prices or {}
    sales_plan_base = sales_plan_base or {}
    sales_fact = sales_fact or {}
    arrival_fixed_dates = arrival_fixed_dates or {}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "План продаж"
    
    # Стили
    thin = Side(style='thin', color='BBBBBB')
    thick = Side(style='medium', color='000000')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    bd_thick_right = Border(left=thin, right=thick, top=thin, bottom=thin)
    
    # СТРОКА 1 - Заголовок
    total_cols = 3 + (SALES_N_MONTHS * 7)  # 3 фикс + 13 месяцев × 7 колонок = 94
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    cell = ws['A1']
    today = datetime.now().strftime("%d.%m.%Y")
    # Автоматический диапазон дат
    start_month_label = get_month_label(0)  # Первый месяц (прошлый)
    end_month_label = get_month_label(SALES_N_MONTHS - 1)  # Последний месяц
    cell.value = f"План продаж  |  {start_month_label} — {end_month_label}  |  {today}"
    cell.font = Font(name='Arial', bold=True, size=11, color='000000')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 20
    
    # СТРОКА 2 - Пустая
    ws.row_dimensions[2].height = 5
    
    # СТРОКА 3 - Заголовки месяцев
    col = 4  # Начинаем с колонки D (после A, B, C)
    for mi in range(SALES_N_MONTHS):
        month_label = get_month_label(mi)
        
        # Объединяем 7 колонок для месяца
        start_col = get_column_letter(col)
        end_col = get_column_letter(col + 6)
        ws.merge_cells(f'{start_col}3:{end_col}3')
        
        cell = ws[f'{start_col}3']
        cell.value = month_label
        cell.font = Font(name='Arial', bold=True, size=10, color='000000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bd
        
        col += 7
    
    ws.row_dimensions[3].height = 20
    
    # СТРОКА 4 - Пустая
    ws.row_dimensions[4].height = 5
    
    # СТРОКА 5 - Подзаголовки колонок
    # Колонка A
    cell = ws['A5']
    cell.value = "Группа / Позиция"
    cell.font = Font(name='Arial', bold=True, size=9, color='000000')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = bd
    ws.column_dimensions['A'].width = 38
    
    # Колонка B
    cell = ws['B5']
    cell.value = "Актуальная\nцена, грн"
    cell.font = Font(name='Arial', bold=True, size=9, color='000000')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = bd
    ws.column_dimensions['B'].width = 11
    
    # Колонка C
    cell = ws['C5']
    cell.value = "План продаж\nбазовый\n(кг/мес)"
    cell.font = Font(name='Arial', bold=True, size=9, color='000000')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = bd
    ws.column_dimensions['C'].width = 11
    
    # Ограничиваем количество групп для теста (определяем ПЕРЕД использованием!)
    groups_to_process = groups[:group_limit] if group_limit else groups
    
    # Подзаголовки для каждого месяца (7 колонок)
    col = 4
    
    # Заголовки БЕЗ недели - неделя будет в строке группы
    headers = [
        "Ост.нач\n(кг)",
        "Приход\n(кг)",
        "План\nпродаж, кг",
        "Факт\nпродаж, кг",
        "%\nвыполнения",
        "План\nпродаж, грн",
        "Факт\nпродаж, грн"
    ]
    
    for mi in range(SALES_N_MONTHS):
        for i, h in enumerate(headers):
            c = ws.cell(row=5, column=col + i)
            c.value = h
            c.font = Font(name='Arial', bold=True, size=8, color='000000')
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Последняя колонка месяца - толстая правая граница
            if i == 6:
                c.border = bd_thick_right
            else:
                c.border = bd
            
            ws.column_dimensions[get_column_letter(col + i)].width = 11
        
        col += 7
    
    ws.row_dimensions[5].height = 35
    
    # Закрепление областей: A, B, C и строки 1-5
    ws.freeze_panes = 'D6'
    
    # ДАННЫЕ ГРУПП
    data_row = 6
    current_month_index = 0  # TODO: определять текущий месяц
    
    for group_idx, group in enumerate(groups_to_process):
        group_name = group["name"]
        group_results = procurement_results.get(group_name, [])
        
        if not group_results:
            continue
        
        # СТРОКА ГРУППЫ
        # Колонка A - название группы
        cell = ws.cell(row=data_row, column=1)
        cell.value = group_name
        cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = bd
        
        # Колонка B - пустая для группы
        cell = ws.cell(row=data_row, column=2)
        cell.value = "—"
        cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bd
        
        # Колонка C - пустая для группы
        cell = ws.cell(row=data_row, column=3)
        cell.value = "—"
        cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bd
        
        # СНАЧАЛА СЧИТАЕМ ИТОГИ ПО ГРУППЕ (для вывода в строке группы)
        group_totals_by_month = {}  # {mi: {plan_kg, fact_kg, plan_uah, fact_uah}}
        
        for mi in range(SALES_N_MONTHS):
            total_plan_kg = 0
            total_fact_kg = 0
            total_plan_uah = 0
            total_fact_uah = 0
            
            # Конвертируем индекс продаж в индекс закупок
            # Продажи: mi=0 (апр), mi=1 (май), mi=2 (июн)...
            # Закупки: mi=0 (май), mi=1 (июн), mi=2 (июл)...
            # Индексы теперь одинаковые
            
            if mi < 0 or mi >= len(group_results):
                # Для прошлого месяца (mi=0, апрель) нет данных в закупках
                group_totals_by_month[mi] = {
                    'plan_kg': 0, 'fact_kg': 0, 'plan_uah': 0, 'fact_uah': 0
                }
                continue
            
            month_data = group_results[mi]
            
            for item_idx, item in enumerate(group["items"]):
                is_seasonal = item.get("seasonal", False)
                item_name = item["name"]
                
                # Цена
                price_data = sales_prices.get(group_idx, {}).get(item_idx, 0)
                if isinstance(price_data, dict):
                    price = price_data.get(mi, 0)
                else:
                    price = price_data
                
                # Базовый план для месяца
                if is_seasonal:
                    month_base_plan = get_plan(item, mi)
                else:
                    base_plan_item = sales_plan_base.get(group_idx, {}).get(item_idx, 0)
                    month_base_plan = base_plan_item
                
                # Данные позиции
                opening_balance = month_data["bsi"].get(item_name, 0)
                arrival_kg = month_data["ia"].get(item_name, 0)
                week_label = month_data.get("wl", "")
                
                # Рабочие дни (с учётом зафиксированных дат)
                working_days = calculate_working_days_for_month(
                    mi, group_name, week_label, arrival_kg, arrival_fixed_dates, group
                )
                
                # План продаж
                sales_plan_kg = calculate_sales_plan(opening_balance, arrival_kg, month_base_plan, working_days)
                
                # Факт продаж
                fact_kg = sales_fact.get(group_idx, {}).get(item_idx, {}).get(mi, 0)
                
                # Суммируем
                total_plan_kg += sales_plan_kg
                total_fact_kg += fact_kg
                if price > 0:
                    total_plan_uah += sales_plan_kg * price
                    total_fact_uah += fact_kg * price
            
            group_totals_by_month[mi] = {
                'plan_kg': total_plan_kg,
                'fact_kg': total_fact_kg,
                'plan_uah': round(total_plan_uah),
                'fact_uah': round(total_fact_uah)
            }
        
        # ТЕПЕРЬ ВЫВОДИМ СТРОКУ ГРУППЫ С ИТОГАМИ
        col = 4
        for mi in range(SALES_N_MONTHS):
            # Конвертируем индекс продаж в индекс закупок
            # Индексы теперь одинаковые
            month_data = group_results[mi] if mi >= 0 and mi < len(group_results) else None
            totals = group_totals_by_month[mi]
            
            if month_data:
                week_label = month_data.get("wl", "")
                arrival_kg = month_data.get("arrive", 0)
                
                # Преобразуем week_label в короткий формат
                week_short = ""
                if week_label:
                    if "Тиж. 1" in week_label or "пред" in week_label.lower():
                        week_short = "нед.1"
                    elif "Тиж. 2" in week_label:
                        week_short = "нед.2"
                    elif "Тиж. 3" in week_label:
                        week_short = "нед.3"
                    elif "Тиж. 4" in week_label:
                        week_short = "нед.4"
            else:
                week_short = ""
                arrival_kg = 0
            
            # Колонка 0: Ост.нач - пусто
            cell = ws.cell(row=data_row, column=col)
            cell.value = ""
            cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            cell.border = bd
            
            # Колонка 1: Приход - неделя или дата
            cell = ws.cell(row=data_row, column=col + 1)
            if arrival_kg > 0:
                # Текущий месяц теперь mi=0 (май)
                is_current_month = (mi == 0)
                
                # Проверяем есть ли зафиксированная дата для этой группы
                if is_current_month and arrival_fixed_dates and group_name in arrival_fixed_dates:
                    # Показываем зафиксированную дату
                    fixed_date = arrival_fixed_dates[group_name]
                    cell.value = fixed_date.strftime("%d.%m")
                else:
                    # Показываем неделю
                    cell.value = week_short if week_short else ""
                
                cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.value = ""
            
            cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            cell.border = bd
            
            # Колонка 2: План продаж, кг - ИТОГО ПО ГРУППЕ
            cell = ws.cell(row=data_row, column=col + 2)
            if totals['plan_kg'] > 0:
                cell.value = totals['plan_kg']
                cell.number_format = '#,##0'
            else:
                cell.value = ""
            cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = bd
            
            # Колонка 3: Факт продаж, кг - ИТОГО ПО ГРУППЕ
            cell = ws.cell(row=data_row, column=col + 3)
            if totals['fact_kg'] > 0:
                cell.value = totals['fact_kg']
                cell.number_format = '#,##0'
            else:
                cell.value = ""
            cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = bd
            
            # Колонка 4: % выполнения - пусто
            cell = ws.cell(row=data_row, column=col + 4)
            cell.value = ""
            cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            cell.border = bd
            
            # Колонка 5: План продаж, грн - ИТОГО ПО ГРУППЕ
            cell = ws.cell(row=data_row, column=col + 5)
            if totals['plan_uah'] > 0:
                cell.value = totals['plan_uah']
                cell.number_format = '#,##0'
            else:
                cell.value = ""
            cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = bd
            
            # Колонка 6: Факт продаж, грн - ИТОГО ПО ГРУППЕ
            cell = ws.cell(row=data_row, column=col + 6)
            if totals['fact_uah'] > 0:
                cell.value = totals['fact_uah']
                cell.number_format = '#,##0'
            else:
                cell.value = ""
            cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = bd_thick_right
            
            col += 7
        
        data_row += 1
        
        # СТРОКИ ПОЗИЦИЙ
        for item_idx, item in enumerate(group["items"]):
            item_name = item["name"]
            is_seasonal = item.get("seasonal", False)
            
            # Получаем данные позиции
            # Цена может быть: float (одна на все месяцы) или dict {month_idx: price}
            price_data = sales_prices.get(group_idx, {}).get(item_idx, 0)
            if isinstance(price_data, dict):
                price = price_data.get(0, 0)  # Берём цену для первого месяца
            else:
                price = price_data  # Это уже число
            
            # Базовый план продаж:
            # Для СЕЗОННЫХ позиций — всегда берём из plan_override (игнорируем импорт)
            # Для обычных позиций — из импорта
            if is_seasonal:
                # Берём сезонный план для первого месяца (mi=0)
                base_plan = get_plan(item, 0)
            else:
                # Берём из импорта
                base_plan = sales_plan_base.get(group_idx, {}).get(item_idx, 0)
            
            # Колонка A - название позиции (с отступом)
            cell = ws.cell(row=data_row, column=1)
            cell.value = f"   {item_name}"
            cell.font = Font(name='Arial', size=9, color='000000')
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = bd
            
            # Колонка B - актуальная цена
            cell = ws.cell(row=data_row, column=2)
            cell.value = round(price) if price > 0 else "—"
            if price > 0:
                cell.number_format = '#,##0'  # БЕЗ копеек
            cell.font = Font(name='Arial', size=9, color='000000')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            cell.border = bd
            
            # Колонка C - базовый план продаж
            cell = ws.cell(row=data_row, column=3)
            cell.value = base_plan if base_plan > 0 else "—"
            if base_plan > 0:
                cell.number_format = '#,##0'
            cell.font = Font(name='Arial', size=9, color='000000')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
            cell.border = bd
            
            # Месячные данные
            col = 4
            for mi in range(SALES_N_MONTHS):
                # Конвертируем индекс продаж в индекс закупок
                # Продажи: mi=0 (апр), mi=1 (май), mi=2 (июн)...
                # Закупки: group_results[0] (май), [1] (июн), [2] (июл)...
                # Индексы теперь одинаковые
                
                if mi < 0 or mi >= len(group_results):
                    # Нет данных - пропускаем
                    for i in range(7):
                        cell = ws.cell(row=data_row, column=col + i)
                        cell.value = "—"
                        cell.font = Font(name='Arial', size=9, color='CCCCCC')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        if i == 6:
                            cell.border = bd_thick_right
                        else:
                            cell.border = bd
                    col += 7
                    continue
                
                # Получаем данные из симуляции закупок
                month_data = group_results[mi]
                
                # Извлекаем данные позиции из результатов симуляции
                opening_balance = month_data["bsi"].get(item_name, 0)
                arrival_kg = month_data["ia"].get(item_name, 0)
                week_label = month_data.get("wl", "")  # "Тиж. 1", "Тиж. 2", и т.д.
                in_transit = month_data.get("in_transit", False)
                
                # 1. Ост.нач (кг)
                cell = ws.cell(row=data_row, column=col)
                cell.value = opening_balance if opening_balance > 0 else "—"
                if opening_balance > 0:
                    cell.number_format = '#,##0'
                cell.font = Font(name='Arial', size=9, color='000000')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = bd
                
                # 2. Приход (кг)
                cell = ws.cell(row=data_row, column=col + 1)
                if arrival_kg > 0:
                    # Просто число кг (неделя уже в заголовке!)
                    cell.value = int(arrival_kg)
                    cell.number_format = '#,##0'
                else:
                    cell.value = "—"
                
                cell.font = Font(name='Arial', size=9, color='000000')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Цвет фона: коричневый для фиксированных, голубой для новых
                if arrival_kg > 0 and in_transit:
                    cell.fill = PatternFill(start_color='5D4037', end_color='5D4037', fill_type='solid')
                    cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
                elif arrival_kg > 0:
                    cell.fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='EEF3FA', end_color='EEF3FA', fill_type='solid')
                
                cell.border = bd
                
                # Рассчитываем рабочие дни с учётом зафиксированных дат
                working_days = calculate_working_days_for_month(
                    mi, group_name, week_label, arrival_kg, arrival_fixed_dates, group
                )
                
                # Определяем базовый план для текущего месяца
                # Для СЕЗОННЫХ позиций — берём plan_override для этого месяца
                # Для обычных — берём из импорта (одинаковый на все месяцы)
                if is_seasonal:
                    month_base_plan = get_plan(item, mi)
                else:
                    month_base_plan = base_plan  # Из импорта
                
                # 3. План продаж, кг
                sales_plan_kg = calculate_sales_plan(opening_balance, arrival_kg, month_base_plan, working_days)
                
                cell = ws.cell(row=data_row, column=col + 2)
                cell.value = sales_plan_kg if sales_plan_kg > 0 else "—"
                if sales_plan_kg > 0:
                    cell.number_format = '#,##0'
                cell.font = Font(name='Arial', size=9, color='000000')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                cell.border = bd
                
                # 4. Факт продаж, кг
                fact_kg = sales_fact.get(group_idx, {}).get(item_idx, {}).get(mi, 0)
                
                cell = ws.cell(row=data_row, column=col + 3)
                cell.value = fact_kg if fact_kg > 0 else "—"
                if fact_kg > 0:
                    cell.number_format = '#,##0'
                cell.font = Font(name='Arial', size=9, color='000000')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = bd
                
                # 5. % выполнения
                cell = ws.cell(row=data_row, column=col + 4)
                if fact_kg > 0 and sales_plan_kg > 0:
                    performance_pct = (fact_kg / sales_plan_kg) * 100
                    cell.value = performance_pct
                    cell.number_format = '0.0"%"'
                    
                    # Цветовое кодирование
                    if performance_pct >= 100:
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Зелёный
                    elif performance_pct >= 90:
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Жёлтый
                    else:
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Красный
                else:
                    cell.value = "—"
                
                cell.font = Font(name='Arial', size=9, color='000000')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = bd
                
                # 6. План продаж, грн
                cell = ws.cell(row=data_row, column=col + 5)
                if sales_plan_kg > 0 and price > 0:
                    plan_uah = sales_plan_kg * price
                    cell.value = round(plan_uah)  # Округляем до целого
                    cell.number_format = '#,##0'  # БЕЗ копеек
                else:
                    cell.value = "—"
                
                cell.font = Font(name='Arial', size=9, color='000000')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = bd
                
                # 7. Факт продаж, грн
                cell = ws.cell(row=data_row, column=col + 6)
                if fact_kg > 0 and price > 0:
                    fact_uah = fact_kg * price
                    cell.value = round(fact_uah)  # Округляем до целого
                    cell.number_format = '#,##0'  # БЕЗ копеек
                else:
                    cell.value = "—"
                
                cell.font = Font(name='Arial', size=9, color='000000')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = bd_thick_right
                
                col += 7
            
            data_row += 1
        
        # НЕ добавляем пустую строку между группами
    
    # СТРОКИ ИТОГОВ
    data_row += 1  # Пустая строка перед итогами
    
    # Строка 1: "1. Собственный импорт, грн"
    row_import = data_row
    cell = ws.cell(row=data_row, column=1)
    cell.value = "1. Собственный импорт, грн"
    cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = bd
    
    # Колонки B, C - пустые
    for c in [2, 3]:
        cell = ws.cell(row=data_row, column=c)
        cell.value = ""
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.border = bd
    
    # Итоги по месяцам
    col = 4
    for mi in range(SALES_N_MONTHS):
        # Считаем суммы по позициям
        total_plan_kg = 0
        total_fact_kg = 0
        total_plan_uah = 0
        total_fact_uah = 0
        
        for group in groups_to_process:
            # Находим настоящий индекс группы в GROUPS
            group_idx = None
            for gi, g in enumerate(GROUPS):
                if g["name"] == group["name"]:
                    group_idx = gi
                    break
            
            if group_idx is None:
                continue
            
            group_name = group["name"]
            group_results = procurement_results.get(group_name, [])
            
            # Конвертируем индекс продаж в индекс закупок
            # Индексы теперь одинаковые
            
            if not group_results or mi < 0 or mi >= len(group_results):
                continue
            
            for item_idx, item in enumerate(group["items"]):
                is_seasonal = item.get("seasonal", False)
                
                # Цена
                price_data = sales_prices.get(group_idx, {}).get(item_idx, 0)
                if isinstance(price_data, dict):
                    price = price_data.get(mi, 0)
                else:
                    price = price_data
                
                # Базовый план для месяца
                if is_seasonal:
                    month_base_plan = get_plan(item, mi)
                else:
                    month_base_plan = sales_plan_base.get(group_idx, {}).get(item_idx, 0)
                
                # Данные месяца
                month_data = group_results[mi]
                item_name = item["name"]
                opening_balance = month_data["bsi"].get(item_name, 0)
                arrival_kg = month_data["ia"].get(item_name, 0)
                week_label = month_data.get("wl", "")
                
                # Рабочие дни (с учётом зафиксированных дат)
                working_days = calculate_working_days_for_month(
                    mi, group_name, week_label, arrival_kg, arrival_fixed_dates, group
                )
                
                # План продаж
                sales_plan_kg = calculate_sales_plan(opening_balance, arrival_kg, month_base_plan, working_days)
                
                # Факт продаж
                fact_kg = sales_fact.get(group_idx, {}).get(item_idx, {}).get(mi, 0)
                
                # Суммируем
                total_plan_kg += sales_plan_kg
                total_fact_kg += fact_kg
                if price > 0:
                    total_plan_uah += sales_plan_kg * price
                    total_fact_uah += fact_kg * price
        
        # 1. Ост.нач - пусто
        cell = ws.cell(row=data_row, column=col)
        cell.value = ""
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.border = bd
        
        # 2. Приход - пусто
        cell = ws.cell(row=data_row, column=col + 1)
        cell.value = ""
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.border = bd
        
        # 3. План продаж, кг
        cell = ws.cell(row=data_row, column=col + 2)
        cell.value = total_plan_kg if total_plan_kg > 0 else "—"
        if total_plan_kg > 0:
            cell.number_format = '#,##0'
        cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bd
        
        # 4. Факт продаж, кг
        cell = ws.cell(row=data_row, column=col + 3)
        cell.value = total_fact_kg if total_fact_kg > 0 else "—"
        if total_fact_kg > 0:
            cell.number_format = '#,##0'
        cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bd
        
        # 5. % выполнения - пусто
        cell = ws.cell(row=data_row, column=col + 4)
        cell.value = ""
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.border = bd
        
        # 6. План продаж, грн
        cell = ws.cell(row=data_row, column=col + 5)
        cell.value = round(total_plan_uah) if total_plan_uah > 0 else "—"
        if total_plan_uah > 0:
            cell.number_format = '#,##0'  # БЕЗ копеек!
        cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bd
        
        # 7. Факт продаж, грн
        cell = ws.cell(row=data_row, column=col + 6)
        cell.value = round(total_fact_uah) if total_fact_uah > 0 else "—"
        if total_fact_uah > 0:
            cell.number_format = '#,##0'  # БЕЗ копеек!
        cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bd_thick_right
        
        col += 7
    
    # СТРОКА 2: "2. Собственное производство, грн"
    data_row += 1
    cell = ws.cell(row=data_row, column=1)
    cell.value = "2. Собственное производство, грн"
    cell.font = Font(name='Arial', bold=True, size=10, color='000000')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = bd
    
    # Колонки B, C - пустые
    for c in [2, 3]:
        cell = ws.cell(row=data_row, column=c)
        cell.value = ""
        cell.border = bd
    
    # Месяцы - только "План продаж, грн" = 1850000
    col = 4
    for mi in range(SALES_N_MONTHS):
        # 1-5: пустые (Ост.нач, Приход, План кг, Факт кг, %)
        for i in range(5):
            cell = ws.cell(row=data_row, column=col + i)
            cell.value = ""
            cell.border = bd
        
        # 6. План продаж, грн = 4850000
        cell = ws.cell(row=data_row, column=col + 5)
        cell.value = 4850000
        cell.number_format = '#,##0'
        cell.font = Font(name='Arial', bold=True, size=9, color='000000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
        cell.border = bd
        
        # 7. Факт продаж, грн - пусто
        cell = ws.cell(row=data_row, column=col + 6)
        cell.value = ""
        cell.border = bd_thick_right
        
        col += 7
    
    # СТРОКА 3: "3. Доп ассортимент, грн"
    data_row += 1
    cell = ws.cell(row=data_row, column=1)
    cell.value = "3. Доп ассортимент, грн"
    cell.font = Font(name='Arial', bold=True, size=10, color='000000')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = bd
    
    # Колонки B, C - пустые
    for c in [2, 3]:
        cell = ws.cell(row=data_row, column=c)
        cell.value = ""
        cell.border = bd
    
    # Месяцы - только "План продаж, грн" = 1900000
    col = 4
    for mi in range(SALES_N_MONTHS):
        # 1-5: пустые
        for i in range(5):
            cell = ws.cell(row=data_row, column=col + i)
            cell.value = ""
            cell.border = bd
        
        # 6. План продаж, грн = 1900000
        cell = ws.cell(row=data_row, column=col + 5)
        cell.value = 1900000
        cell.number_format = '#,##0'
        cell.font = Font(name='Arial', bold=True, size=9, color='000000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
        cell.border = bd
        
        # 7. Факт продаж, грн - пусто
        cell = ws.cell(row=data_row, column=col + 6)
        cell.value = ""
        cell.border = bd_thick_right
        
        col += 7
    
    # СТРОКА 4: "4. Торговая марка, грн"
    data_row += 1
    cell = ws.cell(row=data_row, column=1)
    cell.value = "4. Торговая марка, грн"
    cell.font = Font(name='Arial', bold=True, size=10, color='000000')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = bd
    
    # Колонки B, C - пустые
    for c in [2, 3]:
        cell = ws.cell(row=data_row, column=c)
        cell.value = ""
        cell.border = bd
    
    # Месяцы - только "План продаж, грн" = 2000000
    col = 4
    for mi in range(SALES_N_MONTHS):
        # 1-5: пустые
        for i in range(5):
            cell = ws.cell(row=data_row, column=col + i)
            cell.value = ""
            cell.border = bd
        
        # 6. План продаж, грн = 2000000
        cell = ws.cell(row=data_row, column=col + 5)
        cell.value = 2000000
        cell.number_format = '#,##0'
        cell.font = Font(name='Arial', bold=True, size=9, color='000000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
        cell.border = bd
        
        # 7. Факт продаж, грн - пусто
        cell = ws.cell(row=data_row, column=col + 6)
        cell.value = ""
        cell.border = bd_thick_right
        
        col += 7
    
    # СТРОКА 5: "ВСЕГО РЕАЛИЗАЦИЯ, ГРН"
    data_row += 1
    cell = ws.cell(row=data_row, column=1)
    cell.value = "ВСЕГО РЕАЛИЗАЦИЯ, ГРН"
    cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = bd
    
    # Колонки B, C - пустые
    for c in [2, 3]:
        cell = ws.cell(row=data_row, column=c)
        cell.value = ""
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.border = bd
    
    # Месяцы - суммируем строки 1-4
    col = 4
    for mi in range(SALES_N_MONTHS):
        # 1-5: пустые
        for i in range(5):
            cell = ws.cell(row=data_row, column=col + i)
            cell.value = ""
            cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            cell.border = bd
        
        # 6. План продаж, грн = сумма строк 1-4
        # Получаем значение из строки 1 (импорт)
        import_row_num = row_import
        import_cell_col = col + 5
        import_value = ws.cell(row=import_row_num, column=import_cell_col).value
        if import_value == "—" or import_value is None:
            import_value = 0
        
        total_all = import_value + 4850000 + 1900000 + 2000000
        
        cell = ws.cell(row=data_row, column=col + 5)
        cell.value = round(total_all)
        cell.number_format = '#,##0'
        cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bd
        
        # 7. Факт продаж, грн - пусто
        cell = ws.cell(row=data_row, column=col + 6)
        cell.value = ""
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.border = bd_thick_right
        
        col += 7
    
    # Сохраняем файл
    wb.save(filename)
    print(f"✅ Файл '{filename}' успешно создан!")


if __name__ == "__main__":
    # Тест экспорта (пока без реальных данных)
    print("Тестовый экспорт плана продаж...")
    
    # Заглушка для результатов закупок
    test_results = {}
    
    create_sales_excel(
        procurement_results=test_results,
        groups=GROUPS,
        filename="план_продаж_ТЕСТ.xlsx",
        group_limit=2  # Только 2 группы для теста
    )
