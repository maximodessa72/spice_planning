"""
Модуль загрузки данных сезонности из Excel файла
"""

import openpyxl
from typing import Dict, List


# Маппинг цветов на интенсивность (1-6)
COLOR_TO_INTENSITY = {
    'FF00B050': 1,  # Зелёный - низкая
    'FF92D050': 2,  # Светло-зелёный - умеренная
    'FFFFFF00': 3,  # Жёлтый - средняя
    'FFFF0000': 4,  # Красный - высокая
    'FFFFC000': 5,  # Оранжевый - очень высокая
    'FF00B0F0': 6,  # Голубой - максимальная
}

# Названия цветов для легенды
COLOR_NAMES = {
    1: ("🟢 Зелёный", "низкая интенсивность"),
    2: ("🟣 Светло-зелёный", "умеренная интенсивность"),
    3: ("🟡 Жёлтый", "средняя интенсивность"),
    4: ("🔴 Красный", "высокая интенсивность"),
    5: ("🟠 Оранжевый", "очень высокая интенсивность"),
    6: ("🔵 Голубой", "максимальная интенсивность"),
}


def load_seasonality_from_excel(filepath: str = "сезонность.xlsx") -> Dict[str, List[int]]:
    """
    Загрузить данные сезонности из Excel файла
    
    Returns:
        Dict[название_продукта, список_интенсивностей_по_месяцам]
        Пример: {"Базилік": [0, 0, 0, 3, 0, ...]}
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb['сезонность']
    
    seasonality_data = {}
    
    # Колонки месяцев: D-O (индексы 4-15)
    month_cols = list(range(4, 16))
    
    for row in range(3, ws.max_row + 1):
        # Колонка C (индекс 3) - название продукта
        name_cell = ws.cell(row=row, column=3)
        product_name = name_cell.value
        
        if product_name and str(product_name).strip():
            # Читаем цвета по месяцам
            intensities = []
            has_data = False
            
            for col in month_cols:
                cell = ws.cell(row=row, column=col)
                fill = cell.fill
                
                intensity = 0
                if fill and fill.start_color and hasattr(fill.start_color, 'rgb'):
                    rgb = str(fill.start_color.rgb)
                    if rgb in COLOR_TO_INTENSITY:
                        intensity = COLOR_TO_INTENSITY[rgb]
                        has_data = True
                
                intensities.append(intensity)
            
            # Добавляем только продукты с данными
            if has_data:
                seasonality_data[str(product_name).strip()] = intensities
    
    return seasonality_data


def get_monthly_stats(seasonality_data: Dict[str, List[int]]) -> Dict[str, int]:
    """
    Получить статистику - сколько продуктов имеют пик в каждом месяце
    """
    months = ['ЯНВ', 'ФЕВ', 'МАР', 'АПР', 'МАЙ', 'ИЮН', 'ИЮЛ', 'АВГ', 'СЕН', 'ОКТ', 'НОЯ', 'ДЕК']
    
    monthly_stats = {}
    for month_idx, month_name in enumerate(months):
        count = sum(1 for values in seasonality_data.values() if values[month_idx] > 0)
        monthly_stats[month_name] = count
    
    return monthly_stats


def get_products_by_month(seasonality_data: Dict[str, List[int]], month_idx: int) -> List[tuple]:
    """
    Получить список продуктов с пиками в указанном месяце
    
    Returns:
        List[(название, интенсивность)]
    """
    products = []
    for product_name, intensities in seasonality_data.items():
        if intensities[month_idx] > 0:
            products.append((product_name, intensities[month_idx]))
    
    # Сортируем по интенсивности (убывание)
    return sorted(products, key=lambda x: x[1], reverse=True)


def get_intensity_color(intensity: int) -> str:
    """
    Получить CSS цвет для интенсивности
    """
    colors = {
        0: 'white',
        1: '#00B050',  # Зелёный
        2: '#92D050',  # Светло-зелёный
        3: '#FFFF00',  # Жёлтый
        4: '#FF0000',  # Красный
        5: '#FFC000',  # Оранжевый
        6: '#00B0F0',  # Голубой
    }
    return colors.get(intensity, 'white')


def get_text_color(intensity: int) -> str:
    """
    Получить цвет текста для интенсивности (для контрастности)
    """
    # Белый текст для тёмных фонов
    if intensity in [1, 4, 6]:
        return 'white'
    # Чёрный текст для светлых фонов
    return 'black'
