"""
Веб-приложение для системы планирования закупок
Интерактивный интерфейс для редактирования данных и просмотра результатов
"""

import math
import streamlit as st
import pandas as pd
import json
from data import GROUPS, N_MONTHS
from simulation import (
    run_all_simulations, 
    get_summary_stats, 
    get_critical_groups,
    get_plan
)

# ========== АВТОСОХРАНЕНИЕ POSTGRESQL ==========
def get_db_connection():
    """Подключение к PostgreSQL"""
    try:
        import psycopg2
        from psycopg2.extras import Json
        
        db_url = st.secrets.get("postgres", {}).get("url")
        if not db_url:
            return None
        
        conn = psycopg2.connect(db_url)
        return conn
    except Exception as e:
        st.error(f"Ошибка подключения к БД: {e}")
        return None

def save_state_to_db():
    """Сохранение состояния в PostgreSQL"""
    conn = get_db_connection()
    if not conn:
        return
    
    try:
        cur = conn.cursor()
        
        # Создаём таблицу если нет
        cur.execute("""
            CREATE TABLE IF NOT EXISTS app_state (
                id INTEGER PRIMARY KEY DEFAULT 1,
                state_data JSONB NOT NULL,
                updated_at TIMESTAMP DEFAULT NOW()
            )
        """)
        
        state_data = {
            "groups": st.session_state.groups,
            "sales_plan_base": st.session_state.get("sales_plan_base", {}),
            "sales_prices": st.session_state.get("sales_prices", {}),
            "arrival_fixed_dates": {k: v.isoformat() if hasattr(v, 'isoformat') else str(v) 
                                   for k, v in st.session_state.get("arrival_fixed_dates", {}).items()}
        }
        
        # Upsert
        cur.execute("""
            INSERT INTO app_state (id, state_data, updated_at)
            VALUES (1, %s, NOW())
            ON CONFLICT (id) 
            DO UPDATE SET state_data = %s, updated_at = NOW()
        """, (json.dumps(state_data), json.dumps(state_data)))
        
        conn.commit()
    except Exception as e:
        st.error(f"Ошибка сохранения: {e}")
    finally:
        if conn:
            conn.close()

def load_state_from_db():
    """Загрузка состояния из PostgreSQL"""
    conn = get_db_connection()
    if not conn:
        return None
    
    try:
        cur = conn.cursor()
        cur.execute("SELECT state_data FROM app_state WHERE id = 1")
        result = cur.fetchone()
        
        if result:
            return result[0]
    except Exception as e:
        # Таблица ещё не создана - норм
        pass
    finally:
        if conn:
            conn.close()
    
    return None

# Настройка страницы
st.set_page_config(
    page_title="Система планирования закупок",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)


# ========== АУТЕНТИФИКАЦИЯ ==========
def check_password():
    """Проверка пароля для доступа к приложению с двумя уровнями"""
    
    def password_entered():
        """Проверяет введённый пароль и определяет уровень доступа"""
        entered_password = st.session_state["password"]
        
        # Проверяем admin пароль
        if entered_password == st.secrets["passwords"]["admin"]:
            st.session_state["password_correct"] = True
            st.session_state["user_role"] = "admin"
            del st.session_state["password"]
            return
        
        # Проверяем viewer пароль
        if entered_password == st.secrets["passwords"]["viewer"]:
            st.session_state["password_correct"] = True
            st.session_state["user_role"] = "viewer"
            del st.session_state["password"]
            return
        
        # Проверяем operator пароль
        if entered_password == st.secrets["passwords"].get("operator", ""):
            st.session_state["password_correct"] = True
            st.session_state["user_role"] = "operator"
            del st.session_state["password"]
            return
        
        # Проверяем ved_manager пароль
        if entered_password == st.secrets["passwords"].get("ved_manager", ""):
            st.session_state["password_correct"] = True
            st.session_state["user_role"] = "ved_manager"
            del st.session_state["password"]
            return
        
        # Пароль неверный
        st.session_state["password_correct"] = False
        st.session_state["user_role"] = None

    # Первый запуск - показываем форму входа
    if "password_correct" not in st.session_state:
        st.markdown("### 🔐 Вход в систему планирования")
        st.text_input(
            "Введите пароль:",
            type="password",
            on_change=password_entered,
            key="password"
        )
        st.info("💡 Пароль можно получить у администратора системы")
        return False
    
    # Пароль неверный - показываем ошибку
    elif not st.session_state["password_correct"]:
        st.markdown("### 🔐 Вход в систему планирования")
        st.text_input(
            "Введите пароль:",
            type="password",
            on_change=password_entered,
            key="password"
        )
        st.error("❌ Неверный пароль. Попробуйте ещё раз.")
        return False
    
    # Пароль верный - доступ разрешён
    else:
        return True

# Проверяем пароль перед показом приложения
if not check_password():
    st.stop()  # Останавливаем выполнение если пароль не введён

# ========== КОНЕЦ АУТЕНТИФИКАЦИИ ==========

# Функция очистки старых подтверждённых заказов
def cleanup_old_confirmed_orders():
    """Удаляет подтверждённые заказы из прошлых месяцев (до текущего стартового месяца)"""
    from data import CURRENT_START_MONTH
    
    for group in GROUPS:
        # Очищаем in_transit группы
        old_keys = [mi for mi in group.get("in_transit", {}).keys() if mi < CURRENT_START_MONTH]
        for mi in old_keys:
            del group["in_transit"][mi]
        
        # Очищаем in_transit позиций
        for item in group["items"]:
            old_keys = [mi for mi in item.get("in_transit", {}).keys() if mi < CURRENT_START_MONTH]
            for mi in old_keys:
                del item["in_transit"][mi]

# Инициализация session state
if 'groups' not in st.session_state:
    # Пытаемся загрузить из БД
    loaded_state = load_state_from_db()
    if loaded_state:
        # Обработка как при JSON импорте
        groups = loaded_state.get("groups", GROUPS)
        
        for group in groups:
            if "in_transit" in group:
                group["in_transit"] = {int(k): v for k, v in group["in_transit"].items()}
            if "week_arrival" in group:
                group["week_arrival"] = {int(k): v for k, v in group["week_arrival"].items()}
            
            for item in group.get("items", []):
                if "in_transit" in item:
                    item["in_transit"] = {int(k): v for k, v in item["in_transit"].items()}
                if item.get("seasonal") and "plan_override" in item:
                    po = item["plan_override"]
                    if isinstance(po, dict):
                        item["plan_override"] = [po[str(i)] if str(i) in po else po.get(i, 0) for i in range(12)]
        
        st.session_state.groups = groups
        
        # Sales данные
        sales_plan = loaded_state.get("sales_plan_base", {})
        if sales_plan:
            converted_plan = {}
            for group_key, items in sales_plan.items():
                group_idx = int(group_key)
                converted_plan[group_idx] = {}
                if isinstance(items, dict):
                    for item_key, months in items.items():
                        item_idx = int(item_key)
                        if isinstance(months, dict):
                            converted_plan[group_idx][item_idx] = {int(k): v for k, v in months.items()}
                        else:
                            converted_plan[group_idx][item_idx] = months
            st.session_state.sales_plan_base = converted_plan
        else:
            st.session_state.sales_plan_base = {}
        
        sales_prices = loaded_state.get("sales_prices", {})
        if sales_prices:
            converted_prices = {}
            for group_key, items in sales_prices.items():
                group_idx = int(group_key)
                converted_prices[group_idx] = {}
                if isinstance(items, dict):
                    for item_key, price in items.items():
                        item_idx = int(item_key)
                        converted_prices[group_idx][item_idx] = price
            st.session_state.sales_prices = converted_prices
        else:
            st.session_state.sales_prices = {}
        
        from datetime import datetime
        arrival_dates = loaded_state.get("arrival_fixed_dates", {})
        if arrival_dates:
            converted_dates = {}
            for group_name, date_str in arrival_dates.items():
                try:
                    converted_dates[group_name] = datetime.fromisoformat(date_str)
                except:
                    pass
            st.session_state.arrival_fixed_dates = converted_dates
        else:
            st.session_state.arrival_fixed_dates = {}
    else:
        cleanup_old_confirmed_orders()
        st.session_state.groups = GROUPS
if 'sales_plan_base' not in st.session_state:
    st.session_state.sales_plan_base = {}
if 'sales_prices' not in st.session_state:
    st.session_state.sales_prices = {}
if 'results' not in st.session_state:
    st.session_state.results = None
if 'need_recalc' not in st.session_state:
    st.session_state.need_recalc = True
if 'excel_file' not in st.session_state:
    st.session_state.excel_file = None


def format_number(num):
    """Форматировать число с разделителями тысяч"""
    return f"{num:,}".replace(",", " ")


def get_month_label(mi):
    """Получить название месяца по индексу с учётом текущего стартового месяца"""
    from data import BASE_YEAR, BASE_MONTH, CURRENT_START_MONTH
    
    months = ['Янв','Фев','Мар','Апр','Май','Июн','Июл','Авг','Сен','Окт','Ноя','Дек']
    
    # Добавляем сдвиг от текущего стартового месяца
    absolute_month = CURRENT_START_MONTH + mi
    
    # Вычисляем год и месяц
    y = BASE_YEAR + (BASE_MONTH - 1 + absolute_month) // 12
    m = (BASE_MONTH - 1 + absolute_month) % 12
    
    return f"{months[m]} {y}"


def recalculate():
    """Пересчитать симуляцию"""
    with st.spinner('Пересчёт симуляции...'):
        st.session_state.results = run_all_simulations(st.session_state.groups)
        save_state_to_db()
        st.session_state.need_recalc = False


# Боковое меню
with st.sidebar:
    st.title("📊 Навигация")
    
    # Показываем роль пользователя
    user_role = st.session_state.get("user_role", "viewer")
    if user_role == "admin":
        st.success("👤 Администратор (полный доступ)")
    else:
        st.info("👤 Пользователь (только просмотр)")
    
    st.divider()
    
    # СОХРАНЕНИЕ/ЗАГРУЗКА СОСТОЯНИЯ (только для admin)
    if user_role == "admin":
        st.markdown("### 💾 Резервное копирование")
    
        col1, col2 = st.columns(2)
        
        with col1:
            # Экспорт состояния (доступно всем)
            state_json = json.dumps({
                "groups": st.session_state.groups,
                "sales_plan_base": st.session_state.get("sales_plan_base", {}),
                "sales_prices": st.session_state.get("sales_prices", {}),
                "arrival_fixed_dates": {k: v.isoformat() if hasattr(v, 'isoformat') else str(v) 
                                       for k, v in st.session_state.get("arrival_fixed_dates", {}).items()}
            }, ensure_ascii=False, indent=2)
            
            st.download_button(
                label="📥 Скачать состояние (JSON)",
                data=state_json,
                file_name="app_state_backup.json",
                mime="application/json",
                use_container_width=True,
                help="Скачайте перед обновлением приложения"
            )
        
        with col2:
            # Импорт состояния (доступно всем)
            uploaded_state = st.file_uploader(
                "Загрузить состояние",
                type=['json'],
                key='state_upload',
                label_visibility="collapsed",
                help="Восстановить актуальные данные"
            )
            
            if uploaded_state:
                    try:
                        state_data = json.load(uploaded_state)
                        
                        if st.button("✅ Восстановить состояние", type="primary", use_container_width=True):
                            # Восстанавливаем данные
                            groups = state_data.get("groups", GROUPS)
                            
                            # Исправляем plan_override для сезонных позиций (из dict в list)
                            for group in groups:
                                # Конвертируем ключи in_transit в int
                                if "in_transit" in group:
                                    group["in_transit"] = {int(k): v for k, v in group["in_transit"].items()}
                                if "week_arrival" in group:
                                    group["week_arrival"] = {int(k): v for k, v in group["week_arrival"].items()}
                                
                                for item in group.get("items", []):
                                    # Конвертируем ключи in_transit в int
                                    if "in_transit" in item:
                                        item["in_transit"] = {int(k): v for k, v in item["in_transit"].items()}
                                    
                                    # Исправляем plan_override для сезонных
                                    if item.get("seasonal") and "plan_override" in item:
                                        po = item["plan_override"]
                                        if isinstance(po, dict):
                                            # Конвертируем {0: val, 1: val, ...} в [val, val, ...]
                                            item["plan_override"] = [po[str(i)] if str(i) in po else po.get(i, 0) for i in range(12)]
                            
                            st.session_state.groups = groups
                            
                            # Конвертируем ключи в sales_plan_base {group_idx: {item_idx: {month_idx: value}}}
                            sales_plan = state_data.get("sales_plan_base", {})
                            if sales_plan:
                                converted_plan = {}
                                for group_key, items in sales_plan.items():
                                    group_idx = int(group_key)
                                    converted_plan[group_idx] = {}
                                    if isinstance(items, dict):
                                        for item_key, months in items.items():
                                            item_idx = int(item_key)
                                            if isinstance(months, dict):
                                                converted_plan[group_idx][item_idx] = {int(k): v for k, v in months.items()}
                                            else:
                                                converted_plan[group_idx][item_idx] = months
                                st.session_state.sales_plan_base = converted_plan
                            else:
                                st.session_state.sales_plan_base = {}
                            
                            # Конвертируем ключи в sales_prices {group_idx: {item_idx: price}}
                            sales_prices = state_data.get("sales_prices", {})
                            if sales_prices:
                                converted_prices = {}
                                for group_key, items in sales_prices.items():
                                    group_idx = int(group_key)
                                    converted_prices[group_idx] = {}
                                    if isinstance(items, dict):
                                        for item_key, price in items.items():
                                            item_idx = int(item_key)
                                            converted_prices[group_idx][item_idx] = price
                                st.session_state.sales_prices = converted_prices
                            else:
                                st.session_state.sales_prices = {}
                            
                            # Загружаем даты прихода
                            from datetime import datetime
                            arrival_dates = state_data.get("arrival_fixed_dates", {})
                            if arrival_dates:
                                converted_dates = {}
                                for group_name, date_str in arrival_dates.items():
                                    try:
                                        converted_dates[group_name] = datetime.fromisoformat(date_str)
                                    except:
                                        pass
                                st.session_state.arrival_fixed_dates = converted_dates
                            else:
                                st.session_state.arrival_fixed_dates = {}
                            
                            st.session_state.results = None
                            st.session_state.need_recalc = True
                            
                            st.success("✅ Состояние восстановлено!")
                            st.balloons()
                            st.rerun()
                    except Exception as e:
                        st.error(f"❌ Ошибка загрузки: {e}")
    
    st.divider()
    
    # УРОВЕНЬ 1: Выбор направления
    st.markdown("### 🎯 Направление")
    
    # Для специальных ролей фиксируем направление
    if user_role == "operator":
        direction = "📊 Планирование продаж"
        st.info("👤 Оператор: Планирование продаж")
    elif user_role == "ved_manager":
        direction = "📦 Планирование закупок"
        st.info("👤 Менеджер ВЭД: Планирование закупок")
    else:
        direction = st.radio(
            "Выберите направление:",
            ["📦 Планирование закупок", "📊 Планирование продаж"],
            label_visibility="collapsed"
        )
    
    st.divider()
    
    # УРОВЕНЬ 2: Выбор страницы в зависимости от направления
    if direction == "📦 Планирование закупок":
        st.markdown("### 📄 Страница")
        
        # Для ved_manager показываем только Подтверждение заказов
        if user_role == "ved_manager":
            page = "✅ Подтверждение заказов"
            st.info("👤 Менеджер ВЭД: доступно только подтверждение заказов")
        else:
            page = st.radio(
                "Выберите страницу:",
                [
                    "🏠 Главная",
                    "📥 Импорт данных по закупкам",
                    "✅ Подтверждение заказов",
                    "⚙️ Управление группами",
                    "📊 Редактор данных",
                    "📈 Аналитика по закупкам",
                    "📅 Календарь заказов",
                    "🚚 Календарь поставок",
                    "🌱 Сезоны урожаев"
                ],
                label_visibility="collapsed"
            )
    else:  # Планирование продаж
        st.markdown("### 📄 Страница")
        
        # Для operator показываем только фиксацию дат
        if user_role == "operator":
            page = "✅ Фиксация даты прихода заказа"
            st.info("👤 Оператор: доступна только фиксация дат")
        else:
            page = st.radio(
                "Выберите страницу:",
                [
                    "🏠 Главная (продажи)",
                    "📥 Импорт данных по продажам",
                    "✅ Фиксация даты прихода заказа",
                    "📈 Аналитика по продажам"
                ],
                label_visibility="collapsed"
            )
    
    st.divider()
    
    # Кнопка пересчёта (только для администраторов)
    if user_role == "admin":
        if st.session_state.need_recalc:
            st.warning("⚠️ Есть несохранённые изменения")
        
        if st.button("🔄 Пересчитать", type="primary", use_container_width=True):
            recalculate()
            st.success("✅ Пересчёт завершён!")
            st.rerun()
    
    st.divider()
    
    # Кнопка экспорта (зависит от направления)
    if direction == "📦 Планирование закупок":
        if st.button("💾 Экспорт закупок", use_container_width=True):
            try:
                from excel_export import create_excel
                import tempfile
                import os
                
                # Проверяем что симуляция выполнена
                if st.session_state.results is None:
                    recalculate()
                
                # Создаём файл во временной директории
                with st.spinner('Создание Excel файла...'):
                    # Используем временный файл
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                        output_path = tmp.name
                    
                    create_excel(st.session_state.results, st.session_state.groups, output_path)
                    
                    # Читаем файл в память
                    with open(output_path, 'rb') as f:
                        st.session_state.excel_file = f.read()
                    
                    # Удаляем временный файл
                    try:
                        os.unlink(output_path)
                    except:
                        pass
                
                st.success("✅ Excel файл создан!")
            except Exception as e:
                st.error(f"❌ Ошибка при создании файла: {str(e)}")
                import traceback
                st.code(traceback.format_exc())
        
        # Кнопка скачивания (отдельно, чтобы работала после перезагрузки)
        if st.session_state.excel_file is not None:
            st.download_button(
                label="📥 Скачать план_балансировка.xlsx",
                data=st.session_state.excel_file,
                file_name="план_балансировка.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    
    else:  # Планирование продаж
        # Проверяем что есть импортированные данные
        has_prices = len(st.session_state.get("sales_prices", {})) > 0
        has_plan = len(st.session_state.get("sales_plan_base", {})) > 0
        
        if not has_prices or not has_plan:
            st.warning("⚠️ Сначала импортируйте данные")
            st.caption("Перейдите в раздел 'Импорт данных по продажам'")
        else:
            if st.button("💾 Экспорт продаж", use_container_width=True):
                try:
                    from sales_export import create_sales_excel
                    import tempfile
                    import os
                    
                    # Проверяем что симуляция закупок выполнена
                    if st.session_state.results is None:
                        recalculate()
                    
                    # Создаём файл во временной директории
                    with st.spinner('Создание Excel файла продаж...'):
                        # Используем временный файл
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                            output_path = tmp.name
                        
                        create_sales_excel(
                            procurement_results=st.session_state.results,
                            groups=st.session_state.groups,
                            sales_prices=st.session_state.get("sales_prices", {}),
                            sales_plan_base=st.session_state.get("sales_plan_base", {}),
                            sales_fact=st.session_state.get("sales_fact", {}),
                            arrival_fixed_dates=st.session_state.get("arrival_fixed_dates", {}),
                            filename=output_path
                        )
                        
                        # Читаем файл в память
                        with open(output_path, 'rb') as f:
                            st.session_state.sales_excel_file = f.read()
                        
                        # Удаляем временный файл
                        try:
                            os.unlink(output_path)
                        except:
                            pass
                    
                    st.success("✅ Excel файл продаж создан!")
                except Exception as e:
                    st.error(f"❌ Ошибка при создании файла: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())
            
            # Кнопка скачивания (отдельно, чтобы работала после перезагрузки)
            if st.session_state.get("sales_excel_file") is not None:
                st.download_button(
                    label="📥 Скачать план_продаж.xlsx",
                    data=st.session_state.sales_excel_file,
                    file_name="план_продаж.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )




# ========== ГЛАВНАЯ СТРАНИЦА ==========
if page == "🏠 Главная":
    st.title("📊 Система планирования закупок")
    
    # Показываем текущий горизонт
    from data import CURRENT_START_MONTH
    start_label = get_month_label(0)
    end_label = get_month_label(11)
    st.caption(f"📅 Горизонт планирования: **{start_label} — {end_label}** (12 месяцев)")
    
    if CURRENT_START_MONTH > 0:
        st.info(f"ℹ️ Система автоматически обновила горизонт планирования. Прошлые месяцы удалены, подтверждённые заказы из прошлого очищены.")
    
    # Пересчёт если нужно
    if st.session_state.need_recalc or st.session_state.results is None:
        recalculate()
    
    results = st.session_state.results
    stats = get_summary_stats(results)
    
    # Метрики
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric(
            "Контейнеров",
            format_number(stats["total_containers"]),
            help="Общее количество заказов на 12 месяцев"
        )
    with col2:
        st.metric(
            "Общий объём",
            f"{format_number(stats['total_kg'])} кг",
            help="Общий вес всех заказов на 12 месяцев"
        )
    with col3:
        st.metric(
            "Групп товаров",
            stats["num_groups"],
            help="Количество товарных групп"
        )
    
    st.divider()
    
    # Урожай текущего месяца
    st.subheader("🌱 Урожай текущего месяца")
    
    try:
        import openpyxl
        from datetime import datetime
        
        # Определяем текущий месяц (индекс 0-11)
        current_month_idx = datetime.now().month - 1  # 0=Янв, 1=Фев, ..., 11=Дек
        months_names = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 
                       'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
        months_short = ['ЯНВ', 'ФЕВ', 'МАР', 'АПР', 'МАЙ', 'ИЮН', 'ИЮЛ', 'АВГ', 'СЕН', 'ОКТ', 'НОЯ', 'ДЕК']
        
        # Загружаем данные из Excel
        wb = openpyxl.load_workbook('сезонность.xlsx')
        ws = wb['сезонность']
        
        month_col = 4 + current_month_idx  # Колонка текущего месяца
        
        # Находим все объединённые группы по колонкам B-C (продукты)
        product_groups = {}
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_col <= 3 and merged_range.max_col >= 3:
                rows_in_group = list(range(merged_range.min_row, merged_range.max_row + 1))
                product_groups[merged_range.min_row] = rows_in_group
        
        # Собираем продукты с урожаем
        current_harvest = []
        processed_rows = set()
        
        for row in range(3, ws.max_row + 1):
            if row in processed_rows:
                continue
            
            num_cell = ws.cell(row=row, column=2)
            name_cell = ws.cell(row=row, column=3)
            
            # Пропускаем разделители
            if num_cell.fill and num_cell.fill.start_color and hasattr(num_cell.fill.start_color, 'rgb'):
                rgb = str(num_cell.fill.start_color.rgb)
                if '1F4E79' in rgb:
                    continue
            
            # Пропускаем строки без названия (но не части объединений)
            if not name_cell.value or str(name_cell.value).strip() == '':
                is_part_of_group = False
                for start_row, rows_in_group in product_groups.items():
                    if row in rows_in_group and row != start_row:
                        is_part_of_group = True
                        break
                
                if not is_part_of_group:
                    continue
            
            product_name = str(name_cell.value).strip() if name_cell.value else ""
            
            # Определяем все строки этого продукта
            if row in product_groups:
                rows_to_check = product_groups[row]
                processed_rows.update(rows_to_check)
            else:
                rows_to_check = [row]
                processed_rows.add(row)
            
            # Проверяем все строки на урожай в текущем месяце
            variants = []
            for check_row in rows_to_check:
                comment_cell = ws.cell(row=check_row, column=16)
                comment = str(comment_cell.value).strip() if comment_cell.value else ""
                
                # Проверяем наличие урожая в текущем месяце
                has_harvest_this_month = False
                harvest_color = None
                harvest_period = ""
                
                # Находим все объединённые диапазоны месяцев для этой строки
                for merged in ws.merged_cells.ranges:
                    if merged.min_row == check_row and merged.min_col >= 4 and merged.max_col <= 15:
                        # Это объединение месяцев
                        start_month_idx = merged.min_col - 4  # 0-11
                        end_month_idx = merged.max_col - 4
                        
                        # Текущий месяц попадает в этот диапазон?
                        if start_month_idx <= current_month_idx <= end_month_idx:
                            # Берём цвет из первой ячейки диапазона
                            first_cell = ws.cell(row=check_row, column=merged.min_col)
                            if first_cell.fill and first_cell.fill.start_color and hasattr(first_cell.fill.start_color, 'rgb'):
                                rgb = str(first_cell.fill.start_color.rgb)
                                if rgb and rgb not in ['00000000', 'FFFFFFFF', 'None'] and '1F4E79' not in rgb:
                                    harvest_color = '#' + (rgb[2:8] if rgb.startswith('FF') else rgb[:6])
                                    has_harvest_this_month = True
                                    
                                    # Формируем период
                                    if start_month_idx == end_month_idx:
                                        harvest_period = months_short[start_month_idx]
                                    else:
                                        harvest_period = f"{months_short[start_month_idx]}-{months_short[end_month_idx]}"
                                    break
                
                # Если не нашли объединение - проверяем обычную ячейку текущего месяца
                if not has_harvest_this_month:
                    month_cell = ws.cell(row=check_row, column=month_col)
                    if month_cell.fill and month_cell.fill.start_color and hasattr(month_cell.fill.start_color, 'rgb'):
                        rgb = str(month_cell.fill.start_color.rgb)
                        if rgb and rgb not in ['00000000', 'FFFFFFFF', 'None'] and '1F4E79' not in rgb:
                            harvest_color = '#' + (rgb[2:8] if rgb.startswith('FF') else rgb[:6])
                            has_harvest_this_month = True
                            harvest_period = months_short[current_month_idx]
                
                if has_harvest_this_month and harvest_color:
                    variants.append((comment, harvest_color, harvest_period))
            
            # Добавляем продукт если есть урожай
            if variants and product_name:
                current_harvest.append((product_name, variants))
        
        if current_harvest:
            st.success(f"**{months_names[current_month_idx]}**: созревает **{len(current_harvest)}** продуктов")
            
            # Показываем продукты колонками
            cols = st.columns(3)
            col_idx = 0
            
            for product, variants in current_harvest:
                with cols[col_idx % 3]:
                    if len(variants) == 1:
                        # Один вариант
                        comment, color, period = variants[0]
                        
                        # Формируем текст
                        if comment:
                            display_text = f"{product} ({period}, {comment})"
                        else:
                            display_text = f"{product} ({period})"
                        
                        st.markdown(
                            f'<div style="padding: 5px; margin: 2px; border-left: 4px solid {color};">'
                            f'<span style="font-size: 13px;">• {display_text}</span></div>',
                            unsafe_allow_html=True
                        )
                    else:
                        # Несколько вариантов
                        st.markdown(
                            f'<div style="padding: 5px; margin: 2px; font-weight: bold;">'
                            f'<span style="font-size: 13px;">• {product}</span></div>',
                            unsafe_allow_html=True
                        )
                        for comment, color, period in variants:
                            display_text = f"{comment} ({period})" if comment else period
                            st.markdown(
                                f'<div style="padding: 3px 5px 3px 20px; margin: 1px; border-left: 3px solid {color};">'
                                f'<span style="font-size: 12px; font-style: italic;">→ {display_text}</span></div>',
                                unsafe_allow_html=True
                            )
                
                col_idx += 1
        else:
            st.info(f"В месяце **{months_names[current_month_idx]}** нет продуктов в сезоне урожая")
    
    except FileNotFoundError:
        st.warning("⚠️ Файл 'сезонность.xlsx' не найден - информация об урожае недоступна")
    except Exception as e:
        st.warning(f"⚠️ Не удалось загрузить данные урожая: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
    
    st.divider()
    
    # Критичные группы
    st.subheader("⚠️ Критичные группы (буфер < 1 месяц)")
    critical = get_critical_groups(results, threshold=1.0)
    
    if critical:
        critical_df = pd.DataFrame([
            {
                "Группа": c["group"],
                "Месяц": get_month_label(c["month"]),
                "Буфер (мес)": round(c["buffer"], 2)
            }
            for c in critical[:10]  # Топ-10
        ])
        st.dataframe(critical_df, use_container_width=True, hide_index=True)
    else:
        st.success("✅ Нет критичных групп")


# ============================================================================
# СТРАНИЦА: ИМПОРТ ДАННЫХ
# ============================================================================
elif page == "📥 Импорт данных по закупкам":
    st.title("📥 Импорт данных по закупкам")
    st.info("Обновление остатков и планов закупок")
    
    # Проверка прав доступа
    user_role = st.session_state.get("user_role", "viewer")
    if user_role != "admin":
        st.warning("⚠️ Импорт данных доступен только администраторам")
        st.stop()
    
    st.markdown("""
    ### 📋 Два типа импорта:
    - **Остатки:** Импорт текущих остатков из учётной программы
    - **Планы закупок:** Обновление планов закупок через Excel
    """)
    
    # Вкладки для разных типов импорта
    tab1, tab2 = st.tabs([
        "📦 Остатки из программы",
        "📊 Планы закупок"
    ])
    
    # ========================================================================
    # ВКЛАДКА 1: ИМПОРТ ОСТАТКОВ
    # ========================================================================
    with tab1:
        st.subheader("📦 Импорт остатков из программы")
        st.caption("Загрузите файл с остатками из вашей учётной системы")
        
        uploaded_stocks = st.file_uploader(
            "Выберите файл Excel с остатками",
            type=['xlsx', 'xls'],
            key='stocks_upload',
            help="Файл должен содержать колонки с названиями позиций и остатками"
        )
        
        if uploaded_stocks is not None:
            try:
                # Читаем файл
                df_raw = pd.read_excel(uploaded_stocks, header=None)
                
                # Ищем строку с "Остаток" и "кг"
                data_start = None
                for idx, row in df_raw.iterrows():
                    if pd.notna(row[2]) and 'Остаток' in str(row[2]):
                        data_start = idx + 1
                        break
                
                if data_start:
                    # Извлекаем чистые данные
                    stocks_data = []
                    for idx in range(data_start, len(df_raw)):
                        name = df_raw.iloc[idx, 1]
                        value = df_raw.iloc[idx, 2]
                        
                        if pd.notna(name) and pd.notna(value) and name != 'Итог':
                            stocks_data.append({
                                'Позиция': str(name).strip(),
                                'Остаток': float(value)
                            })
                    
                    if stocks_data:
                        df_stocks = pd.DataFrame(stocks_data)
                        
                        st.success(f"✅ Файл прочитан: {len(df_stocks)} позиций")
                        
                        # Сопоставление с позициями проекта
                        st.subheader("Сопоставление с позициями проекта")
                        
                        matches = []
                        not_found_in_file = []
                        not_found_in_project = []
                        
                        # Создаём словарь для быстрого поиска
                        stocks_dict = {row['Позиция']: row['Остаток'] for _, row in df_stocks.iterrows()}
                        project_positions = set()
                        
                        for group in st.session_state.groups:
                            for item in group['items']:
                                project_positions.add(item['name'])
                                
                                if item['name'] in stocks_dict:
                                    old_balance = item['balance']
                                    new_balance = stocks_dict[item['name']]
                                    
                                    if old_balance != new_balance:
                                        matches.append({
                                            'Группа': group['name'],
                                            'Позиция': item['name'],
                                            'Было': f"{old_balance:,.0f}",
                                            'Станет': f"{new_balance:,.0f}",
                                            'Изменение': f"{new_balance - old_balance:+,.0f}"
                                        })
                                else:
                                    not_found_in_file.append({
                                        'Группа': group['name'],
                                        'Позиция': item['name']
                                    })
                        
                        # Позиции из файла, которых нет в проекте
                        for position in stocks_dict.keys():
                            if position not in project_positions:
                                not_found_in_project.append({
                                    'Позиция': position,
                                    'Остаток': f"{stocks_dict[position]:,.0f}"
                                })
                        
                        # Статистика
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Найдено совпадений", len(matches))
                        with col2:
                            st.metric("Не найдено в файле", len(not_found_in_file))
                        with col3:
                            st.metric("Не найдено в проекте", len(not_found_in_project))
                        
                        # Показываем изменения
                        if matches:
                            st.subheader("Изменения остатков")
                            matches_df = pd.DataFrame(matches)
                            st.dataframe(matches_df, use_container_width=True, hide_index=True)
                            
                            # Кнопка применить
                            st.divider()
                            col1, col2 = st.columns([3, 1])
                            with col1:
                                st.warning("⚠️ Это действие обновит остатки и запустит пересчёт")
                            with col2:
                                if st.button("✅ Применить остатки", type="primary", use_container_width=True, key='apply_stocks'):
                                    # Применяем изменения
                                    for group in st.session_state.groups:
                                        for item in group['items']:
                                            if item['name'] in stocks_dict:
                                                item['balance'] = int(stocks_dict[item['name']])
                                    
                                    # Пересчёт
                                    st.session_state.results = run_all_simulations(st.session_state.groups)
                                    save_state_to_db()
                                    st.session_state.need_recalc = False
                                    
                                    st.success("✅ Остатки обновлены и пересчитаны!")
                                    st.balloons()
                                    st.rerun()
                        else:
                            st.info("ℹ️ Изменений остатков не обнаружено")
                        
                        # Предупреждения
                        if not_found_in_file:
                            with st.expander(f"⚠️ Не найдено в файле ({len(not_found_in_file)} позиций)"):
                                st.dataframe(pd.DataFrame(not_found_in_file), use_container_width=True, hide_index=True)
                        
                        if not_found_in_project:
                            with st.expander(f"⚠️ Не найдено в проекте ({len(not_found_in_project)} позиций)"):
                                st.dataframe(pd.DataFrame(not_found_in_project), use_container_width=True, hide_index=True)
                    
                    else:
                        st.error("❌ Не удалось извлечь данные из файла")
                else:
                    st.error("❌ Не найдена строка с заголовком 'Остаток'. Проверьте формат файла.")
            
            except Exception as e:
                st.error(f"❌ Ошибка при чтении файла: {e}")
    
    # ========================================================================
    # ВКЛАДКА 2: ОБНОВЛЕНИЕ ПЛАНОВ ЗАКУПОК
    # ========================================================================
    with tab2:
        st.subheader("📊 Обновление планов закупок")
        st.caption("Загрузите файл с обновлёнными планами в формате: Название позиции - Количество")
        
        st.markdown("""
        **Формат файла:**
        - Колонка 1: Название позиции
        - Колонка 2: План закупок (кг)
        - Строки с заголовками групп имеют пустое значение во второй колонке
        """)
        
        uploaded_plans = st.file_uploader(
            "Выберите файл Excel с планами закупок",
            type=['xlsx', 'xls'],
            key='plans_upload',
            help="Загрузите файл с обновлёнными планами"
        )
        
        if uploaded_plans is not None:
            try:
                # Читаем файл
                df_raw = pd.read_excel(uploaded_plans, header=None)
                
                st.success(f"✅ Файл загружен: {len(df_raw)} строк")
                
                # Парсим данные
                plan_changes = []
                not_found_positions = []
                
                for idx, row in df_raw.iterrows():
                    # Пропускаем заголовки и пустые строки
                    if idx == 0 or pd.isna(row[0]) or pd.isna(row[1]):
                        continue
                    
                    position_name = str(row[0]).strip()
                    new_plan = row[1]
                    
                    # Пропускаем строки с заголовками групп (NaN во второй колонке)
                    if pd.isna(new_plan):
                        continue
                    
                    # Пропускаем текстовые заголовки
                    try:
                        new_plan = float(new_plan)
                    except (ValueError, TypeError):
                        continue
                    
                    # Ищем позицию в группах
                    found = False
                    for group in st.session_state.groups:
                        for item in group['items']:
                            if item['name'] == position_name:
                                found = True
                                old_plan = item['plan']
                                
                                # Проверяем что позиция не сезонная
                                is_seasonal = item.get('seasonal', False)
                                
                                if not is_seasonal and old_plan != new_plan:
                                    plan_changes.append({
                                        'Группа': group['name'],
                                        'Позиция': position_name,
                                        'Было': f"{old_plan:,.0f}",
                                        'Станет': f"{new_plan:,.0f}",
                                        'Изменение': f"{new_plan - old_plan:+,.0f}"
                                    })
                                elif is_seasonal:
                                    plan_changes.append({
                                        'Группа': group['name'],
                                        'Позиция': position_name + " (сезонная)",
                                        'Было': f"{old_plan:,.0f}",
                                        'Станет': "не меняется",
                                        'Изменение': "—"
                                    })
                                break
                        if found:
                            break
                    
                    if not found:
                        not_found_positions.append({
                            'Позиция': position_name,
                            'План': f"{new_plan:,.0f}"
                        })
                
                # Статистика
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Найдено изменений", len([c for c in plan_changes if c['Станет'] != "не меняется"]))
                with col2:
                    st.metric("Не найдено в проекте", len(not_found_positions))
                
                # Показываем изменения
                if plan_changes:
                    st.subheader("Изменения планов")
                    changes_df = pd.DataFrame(plan_changes)
                    st.dataframe(changes_df, use_container_width=True, hide_index=True)
                    
                    # Применить
                    st.divider()
                    col1, col2 = st.columns([3, 1])
                    with col1:
                        st.warning("⚠️ Это действие обновит планы и запустит пересчёт")
                    with col2:
                        if st.button("✅ Применить планы", type="primary", use_container_width=True, key='apply_plans'):
                            # Применяем изменения
                            for idx, row in df_raw.iterrows():
                                if idx == 0 or pd.isna(row[0]) or pd.isna(row[1]):
                                    continue
                                
                                position_name = str(row[0]).strip()
                                new_plan = float(row[1])
                                
                                # Находим и обновляем позицию
                                for group in st.session_state.groups:
                                    for item in group['items']:
                                        if item['name'] == position_name and not item.get('seasonal', False):
                                            item['plan'] = int(new_plan)
                                            break
                            
                            # Пересчёт
                            st.session_state.results = run_all_simulations(st.session_state.groups)
                            save_state_to_db()
                            st.session_state.need_recalc = False
                            
                            st.success("✅ Планы обновлены и пересчитаны!")
                            st.balloons()
                            st.rerun()
                else:
                    st.info("ℹ️ Изменений планов не обнаружено")
                
                # Предупреждения
                if not_found_positions:
                    with st.expander(f"⚠️ Не найдено в проекте ({len(not_found_positions)} позиций)"):
                        st.dataframe(pd.DataFrame(not_found_positions), use_container_width=True, hide_index=True)
            
            except Exception as e:
                st.error(f"❌ Ошибка при чтении файла: {e}")
                import traceback
                st.code(traceback.format_exc())



# ============================================================================
# СТРАНИЦА: ПОДТВЕРЖДЕНИЕ ЗАКАЗОВ
# ============================================================================
elif page == "✅ Подтверждение заказов":
    st.title("✅ Подтверждение заказов")
    st.info("Внесение точной комплектации подтверждённой производителем")
    
    # Пересчёт если нужно
    if st.session_state.need_recalc or st.session_state.results is None:
        with st.spinner('Пересчёт...'):
            st.session_state.results = run_all_simulations(st.session_state.groups)
            save_state_to_db()
            st.session_state.need_recalc = False
    
    results = st.session_state.results
    
    # Вкладки
    tab1, tab2 = st.tabs(["📝 Новое подтверждение", "📋 Подтверждённые заказы"])
    
    # ========================================================================
    # ВКЛАДКА 1: НОВОЕ ПОДТВЕРЖДЕНИЕ
    # ========================================================================
    with tab1:
        st.subheader("Шаг 1: Выбор группы")
        
        # Группируем заказы по группам
        orders_by_group = {}
        
        for group in st.session_state.groups:
            group_results = results[group["name"]]
            group_orders = []
            
            for mi in range(12):
                r = group_results[mi]
                
                if r["containers"] > 0:
                    arrival_mi = mi + math.ceil(group["cycle"] / 30)
                    is_fixed = arrival_mi < len(group_results) and group_results[arrival_mi]["in_transit"]
                    
                    if not is_fixed:
                        group_orders.append({
                            "Месяц заказа": get_month_label(mi),
                            "Месяц прихода": get_month_label(arrival_mi),
                            "Контейнеров": r["containers"],
                            "Вес (кг)": r["order_kg"],
                            "Срочность": "🔴 Очень срочно" if r["w_buf_before"] < 0.5 else "🟠 Срочно" if r["w_buf_before"] < 0.75 else "🟡 Есть время",
                            "_group": group,
                            "_arrival_mi": arrival_mi,
                            "_order_mi": mi
                        })
            
            if group_orders:
                orders_by_group[group["name"]] = group_orders
        
        if orders_by_group:
            # Выбор группы
            group_names_with_orders = list(orders_by_group.keys())
            selected_group_name = st.selectbox(
                "Выберите группу:",
                group_names_with_orders,
                key="confirm_group_select"
            )
            
            st.divider()
            st.subheader("Шаг 2: Выбор заказа")
            
            # Показываем заказы выбранной группы
            group_orders = orders_by_group[selected_group_name]
            
            st.write(f"**Неподтверждённых заказов в группе:** {len(group_orders)}")
            
            # Таблица заказов
            orders_df = pd.DataFrame([{
                "Месяц заказа": o["Месяц заказа"],
                "Месяц прихода": o["Месяц прихода"],
                "Контейнеров": o["Контейнеров"],
                "Вес (кг)": f"{o['Вес (кг)']:,}",
                "Срочность": o["Срочность"]
            } for o in group_orders])
            
            st.dataframe(orders_df, use_container_width=True, hide_index=True)
            
            # Выбор заказа
            order_options = [
                f"{o['Месяц заказа']} → {o['Месяц прихода']} ({o['Вес (кг)']:,} кг, {o['Контейнеров']} конт.)"
                for o in group_orders
            ]
            
            selected_order_idx = st.selectbox(
                "Выберите заказ для подтверждения:",
                range(len(order_options)),
                format_func=lambda x: order_options[x],
                key="confirm_order_select"
            )
            
            selected_order = group_orders[selected_order_idx]
            group = selected_order["_group"]
            arrival_mi = selected_order["_arrival_mi"]
            
            st.divider()
            st.subheader("Шаг 3: Внесение согласованной комплектации")
            
            st.write(f"**Группа:** {group['name']}")
            st.write(f"**Заказ:** {selected_order['Месяц заказа']} → Приход: {selected_order['Месяц прихода']}")
            
            # Общий вес
            total_weight = st.number_input(
                "Общий вес контейнера от производителя (кг):",
                min_value=0,
                value=selected_order["Вес (кг)"],
                step=100,
                key=f"total_weight_{selected_group_name}_{selected_order_idx}"
            )
            
            # Выбор недели прихода
            week_arrival = st.selectbox(
                "Неделя прихода:",
                options=[1, 2, 3, 4],
                format_func=lambda x: f"Неделя {x}",
                key=f"week_arrival_{selected_group_name}_{selected_order_idx}",
                help="Укажите на какую неделю месяца запланирован приход"
            )
            
            st.divider()
            
            # Таблица комплектации
            st.write("**Комплектация от производителя (кг):**")
            
            # Инициализация данных
            fix_key = f'fix_data_{selected_group_name}_{selected_order_idx}'
            if fix_key not in st.session_state:
                st.session_state[fix_key] = {}
                for item in group["items"]:
                    st.session_state[fix_key][item["name"]] = 0
            
            fix_data = st.session_state[fix_key]
            
            # Таблица для ввода
            for item in group["items"]:
                cols = st.columns([3, 2])
                with cols[0]:
                    st.write(item["name"])
                with cols[1]:
                    fix_data[item["name"]] = st.number_input(
                        "Вес (кг)",
                        min_value=0,
                        value=int(fix_data[item["name"]]),
                        step=25,
                        key=f"item_{selected_group_name}_{selected_order_idx}_{item['name']}",
                        label_visibility="collapsed"
                    )
            
            # Кнопка подтверждения
            st.divider()
            
            col1, col2 = st.columns([3, 1])
            with col1:
                st.warning("⚠️ После подтверждения заказ будет помечен как 'Уже заказан'")
            with col2:
                if st.button("✅ Подтвердить заказ", type="primary", use_container_width=True):
                    # Фиксируем в данных
                    group["in_transit"][arrival_mi] = total_weight
                    
                    # Сохраняем номер недели прихода
                    if "week_arrival" not in group:
                        group["week_arrival"] = {}
                    group["week_arrival"][arrival_mi] = week_arrival
                    
                    for item in group["items"]:
                        if fix_data[item["name"]] > 0:
                            item["in_transit"][arrival_mi] = fix_data[item["name"]]
                        else:
                            item["in_transit"][arrival_mi] = 0
                    
                    # Пересчёт
                    st.session_state.results = run_all_simulations(st.session_state.groups)
                    save_state_to_db()
                    st.session_state.need_recalc = False
                    
                    # Очищаем временные данные
                    if fix_key in st.session_state:
                        del st.session_state[fix_key]
                    
                    st.success("✅ Заказ подтверждён!")
                    st.balloons()
                    st.rerun()
        
        else:
            st.info("ℹ️ Нет неподтверждённых заказов")
    
    # ========================================================================
    # ВКЛАДКА 2: ПОДТВЕРЖДЁННЫЕ ЗАКАЗЫ
    # ========================================================================
    with tab2:
        st.subheader("Подтверждённые заказы")
        
        # Собираем все подтверждённые заказы
        fixed_orders = []
        
        for group in st.session_state.groups:
            for mi, weight in group.get("in_transit", {}).items():
                if weight > 0:
                    # Получаем номер недели если есть
                    week_num = group.get("week_arrival", {}).get(mi, None)
                    
                    # Собираем комплектацию
                    composition = []
                    for item in group["items"]:
                        item_weight = item.get("in_transit", {}).get(mi, 0)
                        if item_weight > 0:
                            composition.append(f"{item['name']}: {item_weight:,} кг")
                    
                    fixed_orders.append({
                        "Группа": group["name"],
                        "Месяц прихода": get_month_label(mi),
                        "Неделя": f"Неделя {week_num}" if week_num else "—",
                        "Вес (кг)": f"{weight:,}",
                        "Позиций": len(composition),
                        "Комплектация": "\n".join(composition),
                        "_group": group,
                        "_mi": mi,
                        "_week": week_num
                    })
        
        if fixed_orders:
            # ФИЛЬТР ПО ГРУППАМ
            all_groups = sorted(list(set([o["Группа"] for o in fixed_orders])))
            
            # Добавляем опцию "Все группы"
            filter_options = ["Все группы"] + all_groups
            
            selected_filter = st.selectbox(
                "Показать заказы группы:",
                options=filter_options,
                key="filter_confirmed_orders"
            )
            
            # Фильтруем заказы
            if selected_filter == "Все группы":
                filtered_orders = fixed_orders
            else:
                filtered_orders = [o for o in fixed_orders if o["Группа"] == selected_filter]
            
            st.write(f"**Показано заказов:** {len(filtered_orders)} из {len(fixed_orders)}")
            
            st.divider()
            
            # Показываем список
            for idx, order in enumerate(filtered_orders):
                week_info = f" — {order['Неделя']}" if order['_week'] else ""
                with st.expander(f"**{order['Группа']}** → {order['Месяц прихода']}{week_info} ({order['Вес (кг)']} кг)"):
                    st.write("**Комплектация:**")
                    st.text(order["Комплектация"])
                    
                    # Четыре колонки для кнопок
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        # Кнопка скачать спецификацию
                        if st.button("📥 Скачать спецификацию", key=f"download_{idx}"):
                            import io
                            import openpyxl
                            from openpyxl.styles import Font, PatternFill, Alignment
                            
                            wb = openpyxl.Workbook()
                            ws = wb.active
                            ws.title = "Спецификация заказа"
                            
                            # Заголовок
                            ws.merge_cells('A1:C1')
                            ws['A1'] = f"Спецификация заказа: {order['Группа']}"
                            ws['A1'].font = Font(bold=True, size=14)
                            ws['A1'].alignment = Alignment(horizontal='center')
                            
                            ws['A2'] = f"Месяц прихода: {order['Месяц прихода']}"
                            ws['A3'] = f"Общий вес: {order['Вес (кг)']} кг"
                            
                            # Таблица
                            ws['A5'] = "Позиция"
                            ws['B5'] = "Вес (кг)"
                            ws['A5'].font = Font(bold=True)
                            ws['B5'].font = Font(bold=True)
                            
                            row = 6
                            for item in order["_group"]["items"]:
                                item_weight = item.get("in_transit", {}).get(order["_mi"], 0)
                                if item_weight > 0:
                                    ws.cell(row=row, column=1, value=item["name"])
                                    ws.cell(row=row, column=2, value=item_weight)
                                    row += 1
                            
                            ws.column_dimensions['A'].width = 50
                            ws.column_dimensions['B'].width = 15
                            
                            buffer = io.BytesIO()
                            wb.save(buffer)
                            buffer.seek(0)
                            
                            st.download_button(
                                label="💾 Скачать Excel",
                                data=buffer,
                                file_name=f"Спецификация_{order['Группа']}_{order['Месяц прихода']}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"download_btn_{idx}"
                            )
                    
                    with col2:
                        # Кнопки смещения заказа
                        st.caption("Сместить заказ:")
                        
                        # Проверяем можно ли сместить назад/вперёд
                        from data import CURRENT_START_MONTH, N_MONTHS
                        can_move_back = order["_mi"] > CURRENT_START_MONTH
                        can_move_forward = order["_mi"] < (N_MONTHS - 1)
                        
                        # Кнопка сместить назад
                        if st.button("⬅️ На месяц назад", 
                                    key=f"move_back_{idx}", 
                                    disabled=not can_move_back,
                                    use_container_width=True,
                                    help="Переместить на предыдущий месяц" if can_move_back else "Нельзя переместить в прошлое"):
                            # Смещаем на месяц назад
                            old_mi = order["_mi"]
                            new_mi = old_mi - 1
                            
                            # Переносим данные группы
                            group_weight = order["_group"]["in_transit"].pop(old_mi)
                            order["_group"]["in_transit"][new_mi] = group_weight
                            
                            # Переносим данные позиций
                            for item in order["_group"]["items"]:
                                if old_mi in item.get("in_transit", {}):
                                    item_weight = item["in_transit"].pop(old_mi)
                                    item["in_transit"][new_mi] = item_weight
                            
                            # Переносим номер недели если есть
                            if "week_arrival" in order["_group"] and old_mi in order["_group"]["week_arrival"]:
                                week_num = order["_group"]["week_arrival"].pop(old_mi)
                                order["_group"]["week_arrival"][new_mi] = week_num
                            
                            # Пересчёт
                            st.session_state.results = run_all_simulations(st.session_state.groups)
                            save_state_to_db()
                            st.session_state.need_recalc = False
                            
                            st.success(f"✅ Заказ перемещён: {order['Месяц прихода']} → {get_month_label(new_mi)}")
                            st.rerun()
                        
                        # Кнопка сместить вперёд
                        if st.button("➡️ На месяц вперёд", 
                                    key=f"move_forward_{idx}",
                                    disabled=not can_move_forward,
                                    use_container_width=True,
                                    help="Переместить на следующий месяц" if can_move_forward else "Достигнут конец горизонта планирования"):
                            # Смещаем на месяц вперёд
                            old_mi = order["_mi"]
                            new_mi = old_mi + 1
                            
                            # Переносим данные группы
                            group_weight = order["_group"]["in_transit"].pop(old_mi)
                            order["_group"]["in_transit"][new_mi] = group_weight
                            
                            # Переносим данные позиций
                            for item in order["_group"]["items"]:
                                if old_mi in item.get("in_transit", {}):
                                    item_weight = item["in_transit"].pop(old_mi)
                                    item["in_transit"][new_mi] = item_weight
                            
                            # Переносим номер недели если есть
                            if "week_arrival" in order["_group"] and old_mi in order["_group"]["week_arrival"]:
                                week_num = order["_group"]["week_arrival"].pop(old_mi)
                                order["_group"]["week_arrival"][new_mi] = week_num
                            
                            # Пересчёт
                            st.session_state.results = run_all_simulations(st.session_state.groups)
                            save_state_to_db()
                            st.session_state.need_recalc = False
                            
                            st.success(f"✅ Заказ перемещён: {order['Месяц прихода']} → {get_month_label(new_mi)}")
                            st.rerun()
                    
                    with col3:
                        # Кнопка редактирования заказа
                        st.caption("Управление:")
                        if st.button("✏️ Редактировать", key=f"edit_{idx}", use_container_width=True):
                            # Сохраняем только имя группы и mi
                            st.session_state.editing_order = {
                                "group_name": order["Группа"],
                                "mi": order["_mi"],
                                "month_label": order["Месяц прихода"],
                                "week": order["_week"]
                            }
                            st.rerun()
                    
                    with col4:
                        # Кнопка удаления заказа
                        st.caption(" ")  # Выравнивание
                        if st.button("🗑️ Удалить заказ", key=f"delete_{idx}", type="secondary", use_container_width=True):
                            # Полное удаление заказа из всех расчётов
                            if order["_mi"] in order["_group"]["in_transit"]:
                                del order["_group"]["in_transit"][order["_mi"]]
                            
                            # Удаляем из позиций
                            for item in order["_group"]["items"]:
                                if order["_mi"] in item.get("in_transit", {}):
                                    del item["in_transit"][order["_mi"]]
                            
                            # Удаляем номер недели если есть
                            if "week_arrival" in order["_group"] and order["_mi"] in order["_group"]["week_arrival"]:
                                del order["_group"]["week_arrival"][order["_mi"]]
                            
                            # Пересчёт
                            st.session_state.results = run_all_simulations(st.session_state.groups)
                            save_state_to_db()
                            st.session_state.need_recalc = False
                            
                            st.success("✅ Заказ полностью удалён!")
                            st.rerun()
        else:
            st.info("ℹ️ Нет подтверждённых заказов")
        
    # МОДАЛЬНОЕ ОКНО РЕДАКТИРОВАНИЯ (ВСЕГДА ДОСТУПНО)
    if "editing_order" in st.session_state:
        st.divider()
        st.subheader(f"✏️ Редактирование заказа: {st.session_state.editing_order['group_name']}")
        st.caption(f"Месяц прихода: {st.session_state.editing_order['month_label']}")
        
        # Находим группу в st.session_state.groups
        group_name = st.session_state.editing_order["group_name"]
        editing_mi = st.session_state.editing_order["mi"]
        
        editing_group = None
        for g in st.session_state.groups:
            if g["name"] == group_name:
                editing_group = g
                break
        
        if not editing_group:
            st.error("❌ Группа не найдена")
            del st.session_state.editing_order
            st.rerun()
        
        st.write("**Текущая комплектация:**")
        
        # Создаём форму редактирования
        with st.form(key="edit_order_form"):
            edited_items = []
            total_edited_weight = 0
            
            for item in editing_group["items"]:
                current_weight = item.get("in_transit", {}).get(editing_mi, 0)
                
                col_name, col_weight = st.columns([3, 1])
                
                with col_name:
                    st.write(f"**{item['name']}**")
                
                with col_weight:
                    new_weight = st.number_input(
                        "кг",
                        min_value=0,
                        value=int(current_weight),
                        step=100,
                        key=f"edit_item_{item['name']}",
                        label_visibility="collapsed"
                    )
                
                edited_items.append({
                    "item": item,
                    "new_weight": new_weight
                })
                total_edited_weight += new_weight
            
            st.divider()
            st.metric("Общий вес после изменений", f"{total_edited_weight:,} кг")
            
            # Редактирование номера недели
            current_week = st.session_state.editing_order.get("week")
            new_week = st.number_input(
                "Номер недели прихода (опционально)",
                min_value=0,
                max_value=4,
                value=current_week if current_week else 0,
                help="0 = не указано, 1-4 = номер недели месяца"
            )
            
            col_save, col_cancel = st.columns(2)
            
            with col_save:
                submit = st.form_submit_button("💾 Сохранить изменения", type="primary", use_container_width=True)
            
            with col_cancel:
                cancel = st.form_submit_button("❌ Отменить", use_container_width=True)
            
            if submit:
                # Находим группу в st.session_state.groups по имени
                group_name = st.session_state.editing_order["group_name"]
                target_group = None
                for g in st.session_state.groups:
                    if g["name"] == group_name:
                        target_group = g
                        break
                
                if target_group:
                    # Применяем изменения к найденной группе
                    target_group["in_transit"][editing_mi] = total_edited_weight
                    
                    for edited in edited_items:
                        # Находим соответствующую позицию в группе
                        for item in target_group["items"]:
                            if item["name"] == edited["item"]["name"]:
                                # ВСЕГДА сохраняем значение (даже 0) чтобы заказ оставался фиксированным
                                if "in_transit" not in item:
                                    item["in_transit"] = {}
                                item["in_transit"][editing_mi] = edited["new_weight"]
                                break
                    
                    # Сохраняем номер недели
                    if new_week > 0:
                        if "week_arrival" not in target_group:
                            target_group["week_arrival"] = {}
                        target_group["week_arrival"][editing_mi] = new_week
                    else:
                        # Удаляем номер недели если 0
                        if "week_arrival" in target_group and editing_mi in target_group["week_arrival"]:
                            del target_group["week_arrival"][editing_mi]
                    
                    # Если общий вес = 0, удаляем весь заказ
                    if total_edited_weight == 0:
                        if editing_mi in target_group["in_transit"]:
                            del target_group["in_transit"][editing_mi]
                        if "week_arrival" in target_group and editing_mi in target_group["week_arrival"]:
                            del target_group["week_arrival"][editing_mi]
                    
                    # Пересчёт
                    st.session_state.results = run_all_simulations(st.session_state.groups)
                    save_state_to_db()
                    st.session_state.need_recalc = False
                    
                    # Закрываем форму редактирования
                    del st.session_state.editing_order
                    
                    st.success("✅ Заказ обновлён!")
                    st.rerun()
                else:
                    st.error("❌ Ошибка: группа не найдена")
                    del st.session_state.editing_order
                    st.rerun()
            
            if cancel:
                # Закрываем форму без изменений
                del st.session_state.editing_order
                st.rerun()



# ============================================================================
# СТРАНИЦА: УПРАВЛЕНИЕ ГРУППАМИ
# ============================================================================
elif page == "⚙️ Управление группами":
    st.title("⚙️ Управление группами")
    st.info("Активация и деактивация товарных групп")
    
    st.markdown("""
    **Деактивированные группы:**
    - Не участвуют в планировании заказов
    - Не показываются в календарях
    - Остатки и планы сохраняются
    - Можно активировать в любой момент
    """)
    
    st.divider()
    
    # Разделяем на активные и неактивные
    active_groups = [g for g in st.session_state.groups if g.get("active", True)]
    inactive_groups = [g for g in st.session_state.groups if not g.get("active", True)]
    
    # Показываем статистику
    col1, col2 = st.columns(2)
    with col1:
        st.metric("Активных групп", len(active_groups))
    with col2:
        st.metric("Деактивированных групп", len(inactive_groups))
    
    st.divider()
    
    # Две колонки: активные и неактивные
    col_active, col_inactive = st.columns(2)
    
    # ========================================================================
    # АКТИВНЫЕ ГРУППЫ
    # ========================================================================
    with col_active:
        st.subheader("✅ Активные группы")
        
        if active_groups:
            for group in active_groups:
                with st.expander(f"**{group['name']}**"):
                    st.write(f"**Позиций:** {len(group['items'])}")
                    st.write(f"**Цикл поставки:** {group['cycle']} дней")
                    st.write(f"**Размер контейнера:** {format_number(group['container'])} кг")
                    
                    # Показываем позиции
                    st.caption("Позиции:")
                    for item in group['items'][:5]:  # Первые 5
                        st.text(f"  • {item['name']}")
                    if len(group['items']) > 5:
                        st.caption(f"  ... и ещё {len(group['items']) - 5}")
                    
                    st.divider()
                    
                    # Кнопка деактивации
                    if st.button(
                        "🔴 Деактивировать группу", 
                        key=f"deactivate_{group['name']}",
                        type="secondary",
                        use_container_width=True
                    ):
                        group["active"] = False
                        st.session_state.need_recalc = True
                        st.success(f"✅ Группа '{group['name']}' деактивирована")
                        st.rerun()
        else:
            st.info("Все группы деактивированы")
    
    # ========================================================================
    # НЕАКТИВНЫЕ ГРУППЫ
    # ========================================================================
    with col_inactive:
        st.subheader("⏸️ Деактивированные группы")
        
        if inactive_groups:
            for group in inactive_groups:
                with st.expander(f"**{group['name']}** (неактивна)"):
                    st.write(f"**Позиций:** {len(group['items'])}")
                    st.write(f"**Цикл поставки:** {group['cycle']} дней")
                    st.write(f"**Размер контейнера:** {format_number(group['container'])} кг")
                    
                    # Показываем позиции
                    st.caption("Позиции:")
                    for item in group['items'][:5]:
                        st.text(f"  • {item['name']}")
                    if len(group['items']) > 5:
                        st.caption(f"  ... и ещё {len(group['items']) - 5}")
                    
                    st.divider()
                    
                    # Кнопка активации
                    if st.button(
                        "✅ Активировать группу", 
                        key=f"activate_{group['name']}",
                        type="primary",
                        use_container_width=True
                    ):
                        group["active"] = True
                        st.session_state.need_recalc = True
                        st.success(f"✅ Группа '{group['name']}' активирована")
                        st.rerun()
        else:
            st.info("Нет деактивированных групп")
    
    # Предупреждение
    if st.session_state.need_recalc:
        st.divider()
        st.warning("⚠️ Есть несохранённые изменения. Нажмите 'Пересчитать' в боковом меню для применения.")


# ========== РЕДАКТОР ДАННЫХ ==========
elif page == "📊 Редактор данных":
    st.title("📊 Редактор данных")
    st.caption("Изменение остатков, планов и фиксированных приходов")
    
    # Проверка прав доступа
    if st.session_state.get("user_role") != "admin":
        st.warning("⚠️ У вас нет прав для редактирования данных")
        st.info("📖 Доступен только режим просмотра. Для редактирования обратитесь к администратору.")
        st.stop()
    
    # Выбор группы
    group_names = [g["name"] for g in st.session_state.groups]
    selected_group_name = st.selectbox("Выберите группу:", group_names)
    
    # Находим группу
    group_idx = group_names.index(selected_group_name)
    group = st.session_state.groups[group_idx]
    
    st.divider()
    
    # Параметры группы (редактируемые)
    st.subheader("Параметры группы")
    
    col1, col2 = st.columns(2)
    with col1:
        new_cycle = st.number_input(
            "Цикл поставки (дней)",
            min_value=1,
            max_value=365,
            value=group['cycle'],
            step=1,
            key=f"cycle_{group_idx}",
            help="Время от размещения заказа до прихода контейнера"
        )
        if new_cycle != group['cycle']:
            group['cycle'] = new_cycle
            st.session_state.need_recalc = True
            st.success(f"✅ Цикл обновлён: {new_cycle} дней")
    
    with col2:
        new_container = st.number_input(
            "Размер контейнера (кг)",
            min_value=1000,
            max_value=100000,
            value=group['container'],
            step=1000,
            key=f"container_{group_idx}",
            help="Вместимость одного контейнера в килограммах"
        )
        if new_container != group['container']:
            group['container'] = new_container
            st.session_state.need_recalc = True
            st.success(f"✅ Размер контейнера обновлён: {format_number(new_container)} кг")
    
    st.divider()
    
    # Редактор позиций
    st.subheader("Позиции в группе")
    
    for item_idx, item in enumerate(group["items"]):
        with st.expander(f"**{item['name']}**", expanded=len(group["items"]) == 1):
            col1, col2 = st.columns(2)
            
            with col1:
                # Остаток на начало
                new_balance = st.number_input(
                    "Остаток на начало (кг)",
                    min_value=0,
                    value=item["balance"],
                    step=100,
                    key=f"balance_{group_idx}_{item_idx}"
                )
                if new_balance != item["balance"]:
                    item["balance"] = new_balance
                    st.session_state.need_recalc = True
                
                # План потребления
                new_plan = st.number_input(
                    "План потребления (кг/мес)",
                    min_value=0,
                    value=item["plan"],
                    step=100,
                    key=f"plan_{group_idx}_{item_idx}"
                )
                if new_plan != item["plan"]:
                    item["plan"] = new_plan
                    st.session_state.need_recalc = True
            
            with col2:
                # Фиксированные приходы
                st.write("**Фиксированные приходы:**")
                
                if item.get("in_transit"):
                    for mi, kg in sorted(item["in_transit"].items()):
                        month = get_month_label(mi)
                        st.write(f"• {month}: {format_number(kg)} кг")
                else:
                    st.write("_Нет фиксированных приходов_")
                
                # TODO: Добавить возможность редактировать фиксированные приходы




# ========== ДЕТАЛЬНЫЙ ПРОСМОТР ==========



# ============================================================================
# СТРАНИЦА: АНАЛИТИКА
# ============================================================================
elif page == "📈 Аналитика по закупкам":
    st.title("📈 Аналитика по закупкам")
    st.caption("Визуализация структуры и динамики заказов")
    
    # Пересчёт если нужно
    if st.session_state.need_recalc or st.session_state.results is None:
        recalculate()
    
    results = st.session_state.results
    
    # Фильтр: показывать ли неактивные группы
    show_inactive = st.checkbox("Показывать неактивные группы", value=False)
    
    st.divider()
    
    # ========================================================================
    # БЛОК 1: СТРУКТУРА ПОСТАВОК
    # ========================================================================
    st.subheader("📊 Структура поставок (12 месяцев)")
    
    # Собираем данные по группам
    group_stats = []
    
    for group in st.session_state.groups:
        is_active = group.get("active", True)
        
        # Пропускаем неактивные если фильтр выключен
        if not is_active and not show_inactive:
            continue
        
        group_results = results[group["name"]]
        total_containers = sum(r["containers"] for r in group_results[:12])
        total_kg = sum(r["order_kg"] for r in group_results[:12])
        
        if total_kg > 0:  # Только если есть заказы
            group_stats.append({
                "Группа": group["name"],
                "Контейнеры": total_containers,
                "Вес (кг)": total_kg,
                "Активна": is_active
            })
    
    if group_stats:
        # Сортируем по весу
        group_stats.sort(key=lambda x: x["Вес (кг)"], reverse=True)
        
        col1, col2 = st.columns(2)
        
        with col1:
            # График 1: Круговая диаграмма по весу (топ-10)
            import plotly.graph_objects as go
            
            top10 = group_stats[:10]
            others_weight = sum(g["Вес (кг)"] for g in group_stats[10:])
            
            labels = [g["Группа"] for g in top10]
            values = [g["Вес (кг)"] for g in top10]
            
            if others_weight > 0:
                labels.append("Остальные")
                values.append(others_weight)
            
            fig = go.Figure(data=[go.Pie(
                labels=labels,
                values=values,
                hole=0.3,
                textinfo='label+percent',
                textposition='auto',
                textfont=dict(size=14),  # Увеличенный шрифт на графике
                hovertemplate='<b>%{label}</b><br>Вес: %{value:,.0f} кг<br>Доля: %{percent}<extra></extra>'
            )])
            
            fig.update_layout(
                title=dict(text="Доля групп по весу", font=dict(size=18, family="Arial")),
                showlegend=False,
                height=400,
                font=dict(size=14, family="Arial")
            )
            
            st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            # График 2: Топ-10 групп по контейнерам
            import plotly.express as px
            
            top10_df = pd.DataFrame(top10)
            
            fig = px.bar(
                top10_df,
                x="Контейнеры",
                y="Группа",
                orientation='h',
                title="Топ-10 групп по контейнерам",
                labels={"Контейнеры": "Контейнеры", "Группа": ""},
                color="Контейнеры",
                color_continuous_scale="Blues"
            )
            
            fig.update_layout(
                showlegend=False,
                height=400,
                xaxis_title="Количество контейнеров",
                yaxis_title="",
                title=dict(font=dict(size=18, family="Arial")),
                font=dict(size=14, family="Arial"),
                xaxis=dict(title_font=dict(size=16)),
                yaxis=dict(categoryorder='total ascending', tickfont=dict(size=14))
            )
            
            fig.update_traces(
                hovertemplate='<b>%{y}</b><br>Контейнеры: %{x}<extra></extra>'
            )
            
            st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("Нет данных для отображения")
    
    st.divider()
    
    # ========================================================================
    # БЛОК 2: ДИНАМИКА ЗАКАЗОВ
    # ========================================================================
    st.subheader("📈 Динамика заказов по месяцам")
    
    # Собираем данные по месяцам
    monthly_data = []
    
    for mi in range(12):
        month_label = get_month_label(mi)
        total_containers = 0
        total_kg = 0
        
        for group in st.session_state.groups:
            is_active = group.get("active", True)
            
            if not is_active and not show_inactive:
                continue
            
            group_results = results[group["name"]]
            if mi < len(group_results):
                r = group_results[mi]
                total_containers += r["containers"]
                total_kg += r["order_kg"]
        
        monthly_data.append({
            "Месяц": month_label,
            "Контейнеры": total_containers,
            "Вес (тонны)": total_kg / 1000
        })
    
    monthly_df = pd.DataFrame(monthly_data)
    
    # Линейный график с двумя осями
    fig = go.Figure()
    
    # Линия контейнеров
    fig.add_trace(go.Scatter(
        x=monthly_df["Месяц"],
        y=monthly_df["Контейнеры"],
        name="Контейнеры",
        line=dict(color='#1f77b4', width=3),
        mode='lines+markers',
        hovertemplate='<b>%{x}</b><br>Контейнеры: %{y}<extra></extra>'
    ))
    
    # Линия веса (вторая ось)
    fig.add_trace(go.Scatter(
        x=monthly_df["Месяц"],
        y=monthly_df["Вес (тонны)"],
        name="Вес (тонны)",
        line=dict(color='#ff7f0e', width=3),
        mode='lines+markers',
        yaxis="y2",
        hovertemplate='<b>%{x}</b><br>Вес: %{y:.1f} т<extra></extra>'
    ))
    
    fig.update_layout(
        title=dict(text="Контейнеры и вес заказов по месяцам", font=dict(size=18, family="Arial")),
        xaxis_title="Месяц",
        yaxis_title="Контейнеры",
        yaxis2=dict(
            title="Вес (тонны)",
            overlaying="y",
            side="right",
            title_font=dict(size=16)
        ),
        hovermode="x unified",
        height=400,
        font=dict(size=14, family="Arial"),
        xaxis=dict(
            title_font=dict(size=16),
            tickfont=dict(size=14)
        ),
        yaxis=dict(
            title_font=dict(size=16),
            tickfont=dict(size=14)
        ),
        legend=dict(font=dict(size=14))
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    st.divider()
    
    # ========================================================================
    # БЛОК 3: БУФЕРЫ И РИСКИ
    # ========================================================================
    st.subheader("⚠️ Буферы и риски")
    
    # Собираем минимальные буферы по группам
    buffer_stats = []
    
    for group in st.session_state.groups:
        is_active = group.get("active", True)
        
        if not is_active and not show_inactive:
            continue
        
        group_results = results[group["name"]]
        
        # Ищем минимальный буфер за 12 месяцев
        min_buffer = 99
        min_month = ""
        
        for r in group_results[:12]:
            if r["w_buf_after"] < min_buffer:
                min_buffer = r["w_buf_after"]
                min_month = get_month_label(r["mi"])
        
        buffer_stats.append({
            "Группа": group["name"],
            "Мин. буфер": min_buffer,
            "Месяц": min_month,
            "Уровень": "Критично" if min_buffer < 1.0 else "Нормально" if min_buffer < 2.0 else "Хорошо"
        })
    
    if buffer_stats:
        col1, col2 = st.columns(2)
        
        with col1:
            # График: Распределение по уровням буфера
            level_counts = {
                "Критично (< 1 мес)": sum(1 for b in buffer_stats if b["Мин. буфер"] < 1.0),
                "Нормально (1-2 мес)": sum(1 for b in buffer_stats if 1.0 <= b["Мин. буфер"] < 2.0),
                "Хорошо (≥ 2 мес)": sum(1 for b in buffer_stats if b["Мин. буфер"] >= 2.0)
            }
            
            # Собираем списки групп для каждого уровня
            level_groups = {
                "Критично (< 1 мес)": [b["Группа"] for b in buffer_stats if b["Мин. буфер"] < 1.0],
                "Нормально (1-2 мес)": [b["Группа"] for b in buffer_stats if 1.0 <= b["Мин. буфер"] < 2.0],
                "Хорошо (≥ 2 мес)": [b["Группа"] for b in buffer_stats if b["Мин. буфер"] >= 2.0]
            }
            
            # Формируем customdata с списками групп
            customdata = []
            for label in level_counts.keys():
                groups_list = level_groups[label]
                groups_text = "<br>".join([f"• {g}" for g in groups_list])
                customdata.append(groups_text)
            
            fig = go.Figure(data=[go.Pie(
                labels=list(level_counts.keys()),
                values=list(level_counts.values()),
                marker=dict(colors=['#ff4444', '#ffaa44', '#44aa44']),
                textinfo='label+value',
                textfont=dict(size=14),
                customdata=customdata,
                hovertemplate='<b>%{label}</b><br>Групп: %{value}<br>Доля: %{percent}<br><br>Группы:<br>%{customdata}<extra></extra>'
            )])
            
            fig.update_layout(
                title=dict(text="Распределение групп по уровню буфера", font=dict(size=18, family="Arial")),
                showlegend=False,
                height=400,
                font=dict(size=14, family="Arial")
            )
            
            st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            # Таблица: Топ-5 критичных групп
            st.write("**Топ-5 самых критичных групп:**")
            
            critical = sorted(buffer_stats, key=lambda x: x["Мин. буфер"])[:5]
            critical_df = pd.DataFrame([{
                "Группа": c["Группа"],
                "Мин. буфер (мес)": f"{c['Мин. буфер']:.2f}",
                "Месяц": c["Месяц"]
            } for c in critical])
            
            st.dataframe(critical_df, use_container_width=True, hide_index=True, height=350)
    
    st.divider()
    
    # ========================================================================
    # БЛОК 4: СТАТИСТИКА ЦИКЛОВ
    # ========================================================================
    st.subheader("⏱️ Распределение по циклам поставки")
    
    # Собираем данные по циклам
    cycle_groups = {
        "14-60 дней": [],
        "60-90 дней": [],
        "90+ дней": []
    }
    
    for group in st.session_state.groups:
        is_active = group.get("active", True)
        
        if not is_active and not show_inactive:
            continue
        
        cycle = group["cycle"]
        
        if cycle <= 60:
            cycle_groups["14-60 дней"].append(group["name"])
        elif cycle <= 90:
            cycle_groups["60-90 дней"].append(group["name"])
        else:
            cycle_groups["90+ дней"].append(group["name"])
    
    # График
    cycle_counts = {k: len(v) for k, v in cycle_groups.items()}
    
    fig = go.Figure(data=[go.Bar(
        x=list(cycle_counts.keys()),
        y=list(cycle_counts.values()),
        marker=dict(color=['#4CAF50', '#FFC107', '#F44336']),
        text=list(cycle_counts.values()),
        textposition='auto',
        textfont=dict(size=16),
        hovertemplate='<b>%{x}</b><br>Групп: %{y}<extra></extra>'
    )])
    
    fig.update_layout(
        title=dict(text="Количество групп по циклам поставки", font=dict(size=18, family="Arial")),
        xaxis_title="Цикл поставки",
        yaxis_title="Количество групп",
        showlegend=False,
        height=400,
        font=dict(size=14, family="Arial"),
        xaxis=dict(
            title_font=dict(size=16),
            tickfont=dict(size=14)
        ),
        yaxis=dict(
            title_font=dict(size=16),
            tickfont=dict(size=14)
        )
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Детализация по клику (expandable)
    with st.expander("Показать детали по циклам"):
        for cycle_name, groups in cycle_groups.items():
            if groups:
                st.write(f"**{cycle_name}** ({len(groups)} групп):")
                for g in groups:
                    st.write(f"• {g}")


# ============================================================================
# СТРАНИЦА: КАЛЕНДАРЬ ЗАКАЗОВ
# ============================================================================
elif page == "📅 Календарь заказов":
    st.title("📅 Календарь заказов")
    st.info("Показывает какие заказы нужно сделать в выбранном месяце")
    
    # Проверяем наличие результатов
    if st.session_state.results is None:
        st.warning("⚠️ Сначала выполните расчёт (кнопка '🔄 Пересчитать' в боковом меню)")
        st.stop()
    
    # Функция определения срочности
    def get_urgency(buffer):
        """Определить срочность по буферу"""
        if buffer < 0.5:
            return "🔴 Очень срочно"
        elif buffer < 0.75:
            return "🟠 Срочно"
        elif buffer < 1.0:
            return "🟡 Есть время"
        else:
            return "🟢 Можно не торопиться"
    
    # ВЫБОР МЕСЯЦА
    month_options = [get_month_label(i) for i in range(12)]  # Первые 12 месяцев
    selected_month = st.selectbox("Выберите месяц:", month_options)
    
    # Находим индекс выбранного месяца
    selected_mi = month_options.index(selected_month)
    
    st.divider()
    
    # Собираем заказы для выбранного месяца
    orders_data = []
    
    for group_name, group_results in st.session_state.results.items():
        r = group_results[selected_mi]
        if r["containers"] > 0:  # Есть заказ в этом месяце
            # Находим группу для получения cycle
            group = next(g for g in st.session_state.groups if g["name"] == group_name)
            arrival_mi = selected_mi + math.ceil(group["cycle"] / 30)
            arrival_month = get_month_label(arrival_mi)
            
            # Определяем срочность
            # ВАЖНО: проверяем in_transit для месяца ПРИБЫТИЯ, а не месяца заказа
            if arrival_mi < len(group_results):
                r_arrival = group_results[arrival_mi]
                if r_arrival["in_transit"]:
                    urgency = "📦 Уже заказан"
                else:
                    urgency = get_urgency(r["w_buf_before"])
            else:
                # Приход за пределами планирования
                urgency = get_urgency(r["w_buf_before"])
            
            orders_data.append({
                "Группа": group_name,
                "Контейнеров": r["containers"],
                "Месяц прибытия": arrival_month,
                "Срочность": urgency
            })
    
    if orders_data:
        orders_df = pd.DataFrame(orders_data)
        
        # Сортируем по срочности
        urgency_order = {
            "🔴 Очень срочно": 0,
            "🟠 Срочно": 1,
            "🟡 Есть время": 2,
            "🟢 Можно не торопиться": 3,
            "📦 Уже заказан": 4
        }
        orders_df["_sort"] = orders_df["Срочность"].map(urgency_order)
        orders_df = orders_df.sort_values("_sort").drop("_sort", axis=1).reset_index(drop=True)
        
        st.subheader(f"Заказы на {selected_month}")
        st.dataframe(orders_df, use_container_width=True, hide_index=True)
        
        # Статистика
        st.divider()
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Групп для заказа", len(orders_df))
        with col2:
            total_containers = orders_df["Контейнеров"].sum()
            st.metric("Всего контейнеров", total_containers)
        with col3:
            # Считаем срочные заказы
            urgent_count = len(orders_df[orders_df["Срочность"].isin(["🔴 Очень срочно", "🟠 Срочно"])])
            st.metric("Срочных заказов", urgent_count)
    else:
        st.info(f"Нет заказов на {selected_month}")


# ============================================================================
# СТРАНИЦА: КАЛЕНДАРЬ ПОСТАВОК
# ============================================================================
elif page == "🚚 Календарь поставок":
    st.title("🚚 Календарь поставок")
    st.info("Показывает когда приходят контейнеры по каждой группе")
    
    # Проверяем наличие результатов
    if st.session_state.results is None:
        st.warning("⚠️ Сначала выполните расчёт (кнопка '🔄 Пересчитать' в боковом меню)")
        st.stop()
    
    # ВЫБОР ПЕРИОДА
    period_options = ["За весь период"] + [get_month_label(i) for i in range(12)]
    selected_period = st.selectbox("Выберите период:", period_options)
    
    st.divider()
    
    # Определяем какие месяцы показывать
    if selected_period == "За весь период":
        months_to_show = list(range(12))
        show_all = True
    else:
        selected_mi = period_options.index(selected_period) - 1  # -1 потому что первый элемент "За весь период"
        months_to_show = [selected_mi]
        show_all = False
    
    # Создаём таблицу: группы × месяцы
    table_data = []
    
    for group in st.session_state.groups:
        group_name = group["name"]
        group_results = st.session_state.results[group_name]
        
        row = {"Группа": group_name}
        
        # Для каждого месяца
        for mi in months_to_show:
            month_label = get_month_label(mi)
            r = group_results[mi]
            
            # Вычисляем количество контейнеров
            if r["arrive"] > 0:
                # Используем unit_container если есть
                container_size = group.get("unit_container", group["container"])
                num_containers = round(r["arrive"] / container_size)
                
                # Добавляем цветной индикатор
                if r["in_transit"]:
                    row[month_label] = f"🟤 {num_containers}"  # Заказан
                else:
                    row[month_label] = f"🟢 {num_containers}"  # Новый
            else:
                row[month_label] = ""
        
        # Итого по группе (только если показываем весь период)
        if show_all:
            total = sum(
                round(group_results[mi]["arrive"] / group.get("unit_container", group["container"])) 
                for mi in months_to_show 
                if group_results[mi]["arrive"] > 0
            )
            row["ВСЕГО"] = total if total > 0 else ""
        
        table_data.append(row)
    
    # Создаём DataFrame
    df = pd.DataFrame(table_data)
    
    # Добавляем строку ИТОГО
    totals_row = {"Группа": "ИТОГО"}
    for mi in months_to_show:
        month_label = get_month_label(mi)
        total = sum(
            round(st.session_state.results[g["name"]][mi]["arrive"] / g.get("unit_container", g["container"]))
            for g in st.session_state.groups
            if st.session_state.results[g["name"]][mi]["arrive"] > 0
        )
        totals_row[month_label] = total if total > 0 else ""
    
    if show_all:
        grand_total = sum(
            round(st.session_state.results[g["name"]][mi]["arrive"] / g.get("unit_container", g["container"]))
            for g in st.session_state.groups
            for mi in months_to_show
            if st.session_state.results[g["name"]][mi]["arrive"] > 0
        )
        totals_row["ВСЕГО"] = grand_total if grand_total > 0 else ""
    
    # Добавляем строку ИТОГО в конец
    df = pd.concat([df, pd.DataFrame([totals_row])], ignore_index=True)
    
    # Отображаем таблицу
    st.subheader(f"Календарь поставок: {selected_period}")
    st.dataframe(df, use_container_width=True, hide_index=True, height=600)
    
    # Статистика
    if show_all:
        st.divider()
        col1, col2 = st.columns(2)
        with col1:
            groups_with_arrivals = len([g for g in st.session_state.groups if any(
                st.session_state.results[g["name"]][mi]["arrive"] > 0 for mi in months_to_show
            )])
            st.metric("Групп с приходами", groups_with_arrivals)
        with col2:
            st.metric("Всего контейнеров", grand_total if 'grand_total' in locals() else 0)
        
    else:
        st.info("Нет приходов для отображения")



# ========== СЕЗОНЫ УРОЖАЕВ ==========
elif page == "🌱 Сезоны урожаев":
    st.title("🌱 Сезоны урожаев")
    st.caption("Календарь созревания продуктов по месяцам года")
    
    # Загружаем данные из Excel
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook('сезонность.xlsx')
        ws = wb['сезонность']
        
        # Находим все объединённые группы по колонкам B-C (продукты)
        product_groups = {}
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_col <= 3 and merged_range.max_col >= 3:  # Колонка C (название)
                rows_in_group = list(range(merged_range.min_row, merged_range.max_row + 1))
                product_groups[merged_range.min_row] = rows_in_group
        
        # Читаем данные
        table_rows = []
        processed_rows = set()
        
        for row in range(3, ws.max_row + 1):
            if row in processed_rows:
                continue
            
            num_cell = ws.cell(row=row, column=2)
            name_cell = ws.cell(row=row, column=3)
            
            # Пропускаем разделители
            is_separator = False
            if num_cell.fill and num_cell.fill.start_color and hasattr(num_cell.fill.start_color, 'rgb'):
                rgb = str(num_cell.fill.start_color.rgb)
                if '1F4E79' in rgb:
                    is_separator = True
            
            if is_separator:
                table_rows.append({'type': 'separator'})
                continue
            
            if not name_cell.value or str(name_cell.value).strip() == '':
                continue
            
            num_val = str(num_cell.value) if num_cell.value else ''
            name_val = str(name_cell.value).strip()
            
            # Многострочный или однострочный продукт
            if row in product_groups:
                rows_in_product = product_groups[row]
                processed_rows.update(rows_in_product)
                
                # Заголовок продукта
                table_rows.append({
                    'type': 'product_header',
                    'num': num_val,
                    'name': name_val,
                })
                
                # Подстроки для каждой страны
                for sub_row in rows_in_product:
                    comment_cell = ws.cell(row=sub_row, column=16)
                    comment = str(comment_cell.value).strip() if comment_cell.value else ''
                    
                    if not comment:
                        continue
                    
                    # Находим объединённые ячейки месяцев для этой строки
                    month_merges = {}
                    for merged in ws.merged_cells.ranges:
                        if merged.min_row == sub_row and merged.min_col >= 4 and merged.max_col <= 15:
                            # Диапазон месяцев (относительно колонки 4)
                            start_month = merged.min_col - 4
                            end_month = merged.max_col - 4
                            for m in range(start_month, end_month + 1):
                                month_merges[m] = (start_month, end_month)
                    
                    # Читаем месяцы с учётом объединений
                    months_data = []
                    has_color = False
                    
                    for month_idx in range(12):
                        col = 4 + month_idx
                        cell = ws.cell(row=sub_row, column=col)
                        
                        # Проверяем объединение
                        if month_idx in month_merges:
                            start_m, end_m = month_merges[month_idx]
                            if month_idx == start_m:
                                # Начало объединения - берём цвет и colspan
                                color = None
                                if cell.fill and cell.fill.start_color and hasattr(cell.fill.start_color, 'rgb'):
                                    rgb = str(cell.fill.start_color.rgb)
                                    if rgb and rgb not in ['00000000', 'FFFFFFFF', 'None'] and '1F4E79' not in rgb:
                                        color = '#' + (rgb[2:8] if rgb.startswith('FF') else rgb[:6])
                                        has_color = True
                                
                                months_data.append({
                                    'color': color,
                                    'colspan': end_m - start_m + 1
                                })
                            else:
                                # Продолжение объединения - пропускаем
                                months_data.append({'colspan': 0})
                        else:
                            # Обычная ячейка
                            color = None
                            if cell.fill and cell.fill.start_color and hasattr(cell.fill.start_color, 'rgb'):
                                rgb = str(cell.fill.start_color.rgb)
                                if rgb and rgb not in ['00000000', 'FFFFFFFF', 'None'] and '1F4E79' not in rgb:
                                    color = '#' + (rgb[2:8] if rgb.startswith('FF') else rgb[:6])
                                    has_color = True
                            
                            months_data.append({
                                'color': color,
                                'colspan': 1
                            })
                    
                    if has_color:
                        table_rows.append({
                            'type': 'product_subrow',
                            'comment': comment,
                            'months': months_data
                        })
            else:
                # Однострочный продукт
                processed_rows.add(row)
                comment_cell = ws.cell(row=row, column=16)
                comment = str(comment_cell.value).strip() if comment_cell.value else ''
                
                # Находим объединённые ячейки месяцев
                month_merges = {}
                for merged in ws.merged_cells.ranges:
                    if merged.min_row == row and merged.min_col >= 4 and merged.max_col <= 15:
                        start_month = merged.min_col - 4
                        end_month = merged.max_col - 4
                        for m in range(start_month, end_month + 1):
                            month_merges[m] = (start_month, end_month)
                
                # Читаем месяцы
                months_data = []
                has_color = False
                
                for month_idx in range(12):
                    col = 4 + month_idx
                    cell = ws.cell(row=row, column=col)
                    
                    if month_idx in month_merges:
                        start_m, end_m = month_merges[month_idx]
                        if month_idx == start_m:
                            color = None
                            if cell.fill and cell.fill.start_color and hasattr(cell.fill.start_color, 'rgb'):
                                rgb = str(cell.fill.start_color.rgb)
                                if rgb and rgb not in ['00000000', 'FFFFFFFF', 'None'] and '1F4E79' not in rgb:
                                    color = '#' + (rgb[2:8] if rgb.startswith('FF') else rgb[:6])
                                    has_color = True
                            
                            months_data.append({
                                'color': color,
                                'colspan': end_m - start_m + 1
                            })
                        else:
                            months_data.append({'colspan': 0})
                    else:
                        color = None
                        if cell.fill and cell.fill.start_color and hasattr(cell.fill.start_color, 'rgb'):
                            rgb = str(cell.fill.start_color.rgb)
                            if rgb and rgb not in ['00000000', 'FFFFFFFF', 'None'] and '1F4E79' not in rgb:
                                color = '#' + (rgb[2:8] if rgb.startswith('FF') else rgb[:6])
                                has_color = True
                        
                        months_data.append({
                            'color': color,
                            'colspan': 1
                        })
                
                if has_color:
                    table_rows.append({
                        'type': 'product_single',
                        'num': num_val,
                        'name': name_val,
                        'comment': comment,
                        'months': months_data
                    })
        
        if not table_rows:
            st.warning("⚠️ Данные сезонности не найдены в файле")
        else:
            months = ['ЯНВ', 'ФЕВ', 'МАР', 'АПР', 'МАЙ', 'ИЮН', 'ИЮЛ', 'АВГ', 'СЕН', 'ОКТ', 'НОЯ', 'ДЕК']
            
            count_products = len([r for r in table_rows if r['type'] in ['product_header', 'product_single']])
            st.subheader("📊 Календарь урожая")
            st.caption(f"Показано {count_products} продуктов")
            
            # HTML таблица
            table_html = '''
            <style>
                .seasonality-table-container {
                    max-height: 700px;
                    overflow-y: auto;
                    border: 1px solid #ddd;
                    border-radius: 4px;
                }
                .seasonality-table {
                    width: 100%;
                    border-collapse: collapse;
                    font-size: 13px;
                }
                .seasonality-table thead {
                    position: sticky;
                    top: 0;
                    z-index: 10;
                    background-color: #1F4E79;
                    color: white;
                }
                .seasonality-table th {
                    border: 1px solid #ddd;
                    padding: 10px 8px;
                    text-align: center;
                    font-weight: bold;
                }
                .seasonality-table td {
                    border: 1px solid #ddd;
                    padding: 8px;
                }
                .seasonality-table tr.product-row:hover {
                    background-color: #f9f9f9;
                }
                .seasonality-table tr.separator-row {
                    height: 8px;
                    background-color: #f0f0f0;
                }
                .seasonality-table tr.separator-row td {
                    border: none;
                    padding: 0;
                }
                .seasonality-table tr.subrow {
                    background-color: #fafafa;
                }
                .seasonality-table tr.header-row {
                    font-weight: bold;
                }
            </style>
            <div class="seasonality-table-container">
                <table class="seasonality-table">
                    <thead>
                        <tr>
                            <th style="text-align: left; width: 40px;">№</th>
                            <th style="text-align: left; width: 300px;">Продукт</th>
            '''
            
            for month in months:
                table_html += f'<th style="width: 65px;">{month}</th>'
            table_html += '</tr></thead><tbody>'
            
            # Строки
            for row_data in table_rows:
                if row_data['type'] == 'separator':
                    table_html += f'<tr class="separator-row"><td colspan="14"></td></tr>'
                
                elif row_data['type'] == 'product_header':
                    table_html += '<tr class="product-row header-row">'
                    table_html += f'<td style="text-align: center; color: #666;">{row_data["num"]}</td>'
                    table_html += f'<td style="font-weight: bold;">{row_data["name"]}</td>'
                    for _ in range(12):
                        table_html += '<td></td>'
                    table_html += '</tr>'
                
                elif row_data['type'] == 'product_subrow':
                    table_html += '<tr class="product-row subrow">'
                    table_html += '<td></td>'
                    table_html += f'<td style="padding-left: 25px; font-style: italic; color: #555;">→ {row_data["comment"]}</td>'
                    
                    for month_data in row_data['months']:
                        if month_data['colspan'] == 0:
                            continue
                        
                        color = month_data.get('color')
                        colspan = month_data['colspan']
                        
                        if colspan > 1:
                            bg = f'background-color: {color};' if color else ''
                            table_html += f'<td colspan="{colspan}" style="{bg}"></td>'
                        else:
                            bg = f'background-color: {color};' if color else ''
                            table_html += f'<td style="{bg}"></td>'
                    
                    table_html += '</tr>'
                
                elif row_data['type'] == 'product_single':
                    table_html += '<tr class="product-row">'
                    table_html += f'<td style="text-align: center; color: #666;">{row_data["num"]}</td>'
                    table_html += f'<td style="font-weight: bold;">{row_data["name"]}</td>'
                    
                    for month_data in row_data['months']:
                        if month_data['colspan'] == 0:
                            continue
                        
                        color = month_data.get('color')
                        colspan = month_data['colspan']
                        
                        if colspan > 1:
                            bg = f'background-color: {color};' if color else ''
                            table_html += f'<td colspan="{colspan}" style="{bg}"></td>'
                        else:
                            bg = f'background-color: {color};' if color else ''
                            table_html += f'<td style="{bg}"></td>'
                    
                    table_html += '</tr>'
            
            table_html += '</tbody></table></div>'
            
            st.markdown(table_html, unsafe_allow_html=True)
    
    except FileNotFoundError:
        st.error("❌ Файл 'сезонность.xlsx' не найден")
        st.info("Убедитесь что файл находится в той же папке что и приложение")
    except Exception as e:
        st.error(f"❌ Ошибка при загрузке данных: {str(e)}")
        import traceback
        st.code(traceback.format_exc())


# ========== СТРАНИЦЫ ПЛАНИРОВАНИЯ ПРОДАЖ ==========

elif page == "🏠 Главная (продажи)":
    st.title("📊 Планирование продаж")
    st.caption("📅 Управление планами продаж и аналитика")
    
    st.info("🚧 **В разработке**")
    
    st.markdown("""
    ### 📋 Возможности модуля:
    
    - **Импорт данных** — загрузка актуальных цен, планов и фактов продаж
    - **Фиксация прихода** — учёт дат поступления заказов для корректировки планов
    - **Аналитика** — визуализация динамики продаж и отклонений от плана
    - **Экспорт** — выгрузка отчётов в Excel
    
    Выберите раздел в меню слева для начала работы.
    """)


elif page == "📥 Импорт данных по продажам":
    st.title("📥 Импорт данных по продажам")
    st.info("Загрузка актуальных цен, планов и фактов продаж")
    
    # Проверка прав доступа
    user_role = st.session_state.get("user_role", "viewer")
    if user_role != "admin":
        st.warning("⚠️ Импорт данных доступен только администраторам")
        st.stop()
    
    # Инициализация данных продаж в session_state
    if "sales_prices" not in st.session_state:
        st.session_state.sales_prices = {}
    if "sales_plan_base" not in st.session_state:
        st.session_state.sales_plan_base = {}
    if "sales_fact" not in st.session_state:
        st.session_state.sales_fact = {}
    
    st.markdown("""
    ### 📋 Три типа импорта данных:
    
    Каждый месяц **первого числа** необходимо загрузить три типа данных:
    """)
    
    # Вкладки для разных типов импорта
    tab1, tab2, tab3 = st.tabs([
        "💰 Актуальная цена",
        "📊 План продаж базовый", 
        "📈 Факт продаж"
    ])
    
    with tab1:
        st.subheader("💰 Импорт актуальной цены")
        st.caption("Загрузка цен на текущий месяц — **первого числа текущего месяца**")
        
        uploaded_prices = st.file_uploader(
            "Загрузите файл с актуальными ценами (Excel)",
            type=['xlsx', 'xls'],
            key="upload_prices"
        )
        
        if uploaded_prices:
            try:
                import tempfile
                import os
                from sales_data import import_prices_from_excel
                
                # Сохраняем во временный файл
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    tmp.write(uploaded_prices.getvalue())
                    tmp_path = tmp.name
                
                # Импортируем цены (НЕ сохраняем в session_state сразу)
                with st.spinner('Загрузка цен...'):
                    prices = import_prices_from_excel(tmp_path)
                    # Временно сохраняем для предпросмотра
                    st.session_state.temp_prices = prices
                
                # Удаляем временный файл
                try:
                    os.unlink(tmp_path)
                except:
                    pass
                
                # Статистика
                total_items = sum(len(items) for items in prices.values())
                total_groups = len(prices)
                
                st.success(f"✅ Файл прочитан: **{total_items}** позиций из **{total_groups}** групп")
                
                # Показываем примеры
                with st.expander("📋 Просмотр загруженных цен (первые 10)"):
                    count = 0
                    for group_idx, items in prices.items():
                        if count >= 10:
                            break
                        group_name = GROUPS[group_idx]["name"]
                        st.markdown(f"**{group_name}**")
                        for item_idx, price in items.items():
                            if count >= 10:
                                break
                            item_name = GROUPS[group_idx]["items"][item_idx]["name"]
                            st.write(f"  • {item_name}: {price} грн")
                            count += 1
                
                # Кнопка применить
                st.divider()
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.warning("⚠️ Это действие обновит актуальные цены")
                with col2:
                    if st.button("✅ Применить цены", type="primary", use_container_width=True, key='apply_prices'):
                        st.session_state.sales_prices = st.session_state.temp_prices
                        del st.session_state.temp_prices
                        st.success("✅ Цены обновлены!")
                        st.balloons()
                        st.rerun()
                
            except Exception as e:
                st.error(f"❌ Ошибка при импорте: {str(e)}")
                import traceback
                st.code(traceback.format_exc())
    
    with tab2:
        st.subheader("📊 Импорт плана продаж базового")
        st.caption("Загрузка базового плана на текущий месяц — **первого числа текущего месяца**")
        
        uploaded_plan = st.file_uploader(
            "Загрузите файл с планом продаж (Excel)",
            type=['xlsx', 'xls'],
            key="upload_plan"
        )
        
        if uploaded_plan:
            try:
                import tempfile
                import os
                from sales_data import import_sales_plan_from_excel
                
                # Сохраняем во временный файл
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    tmp.write(uploaded_plan.getvalue())
                    tmp_path = tmp.name
                
                # Импортируем план (НЕ сохраняем в session_state сразу)
                with st.spinner('Загрузка плана продаж...'):
                    plan = import_sales_plan_from_excel(tmp_path)
                    # Временно сохраняем для предпросмотра
                    st.session_state.temp_plan = plan
                
                # Удаляем временный файл
                try:
                    os.unlink(tmp_path)
                except:
                    pass
                
                # Статистика
                total_items = sum(len(items) for items in plan.values())
                total_groups = len(plan)
                total_kg = sum(sum(items.values()) for items in plan.values())
                
                st.success(f"✅ Файл прочитан: **{total_items}** позиций из **{total_groups}** групп")
                st.info(f"📊 Общий план продаж: **{total_kg:,}** кг/мес")
                
                # Показываем примеры
                with st.expander("📋 Просмотр загруженного плана (первые 10)"):
                    count = 0
                    for group_idx, items in plan.items():
                        if count >= 10:
                            break
                        group_name = GROUPS[group_idx]["name"]
                        st.markdown(f"**{group_name}**")
                        for item_idx, plan_kg in items.items():
                            if count >= 10:
                                break
                            item_name = GROUPS[group_idx]["items"][item_idx]["name"]
                            st.write(f"  • {item_name}: {plan_kg:,} кг")
                            count += 1
                
                # Кнопка применить
                st.divider()
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.warning("⚠️ Это действие обновит план продаж")
                with col2:
                    if st.button("✅ Применить план", type="primary", use_container_width=True, key='apply_plan'):
                        st.session_state.sales_plan_base = st.session_state.temp_plan
                        del st.session_state.temp_plan
                        st.success("✅ План продаж обновлён!")
                        st.balloons()
                        st.rerun()
                
            except Exception as e:
                st.error(f"❌ Ошибка при импорте: {str(e)}")
                import traceback
                st.code(traceback.format_exc())
    
    with tab3:
        st.subheader("📈 Импорт факта продаж")
        st.caption("Загрузка факта по **закрытому месяцу** — первого числа **следующего** месяца")
        
        st.warning("⚠️ **Важно:** Факт продаж загружается в прошлый месяц (месяц -1)")
        
        st.info("🚧 **В разработке** — будет доступна загрузка Excel с фактом продаж")
        
        uploaded_fact = st.file_uploader(
            "Загрузите файл с фактом продаж (Excel)",
            type=['xlsx', 'xls'],
            key="upload_fact",
            disabled=True
        )
        
        if uploaded_fact:
            st.success("✅ Файл загружен")


elif page == "✅ Фиксация даты прихода заказа":
    st.title("✅ Фиксация даты прихода заказа")
    st.info("Фиксация точных дат прихода контейнеров для текущего месяца")
    
    # Проверка прав доступа
    user_role = st.session_state.get("user_role", "viewer")
    if user_role not in ["admin", "operator"]:
        st.warning("⚠️ Фиксация дат доступна только администраторам и операторам")
        st.stop()
    
    # Проверяем что симуляция выполнена (для operator делаем автоматически)
    if st.session_state.results is None:
        if user_role == "operator":
            with st.spinner('Загрузка данных...'):
                st.session_state.results = run_all_simulations(st.session_state.groups)
                save_state_to_db()
                st.session_state.need_recalc = False
        else:
            st.warning("⚠️ Сначала выполните пересчёт закупок")
            st.stop()
    
    # Инициализация хранилища дат
    if "arrival_fixed_dates" not in st.session_state:
        st.session_state.arrival_fixed_dates = {}  # {group_name: datetime}
    
    from datetime import datetime
    from sales_data import SALES_BASE_YEAR, SALES_BASE_MONTH
    
    # Определяем текущий месяц
    # Теперь и закупки, и продажи начинаются с ТЕКУЩЕГО месяца (май)
    # mi=0 = май для обоих
    
    procurement_current_mi = 0  # Текущий месяц (май)
    sales_current_mi = 0  # Текущий месяц (май)
    
    # Вычисляем год и месяц текущего месяца
    current_year = SALES_BASE_YEAR + (SALES_BASE_MONTH - 1 + sales_current_mi) // 12
    current_month = (SALES_BASE_MONTH - 1 + sales_current_mi) % 12 + 1
    
    month_names = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 
                   'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
    current_month_name = month_names[current_month - 1]
    
    st.markdown(f"### 📅 Текущий месяц: **{current_month_name} {current_year}**")
    
    st.markdown("""
    **Автоматические даты по неделям:**
    - `пред мес` / `нед 1` → **5-е число**
    - `нед 2` → **10-е число**
    - `нед 3` → **15-е число**
    - `нед 4` → **20-е число**
    """)
    
    # Собираем заказы текущего месяца
    orders_current_month = []
    
    for group in st.session_state.groups:
        group_name = group["name"]
        group_results = st.session_state.results.get(group_name, [])
        
        # Читаем ТЕКУЩИЙ месяц из симуляции закупок (mi=0)
        if procurement_current_mi < len(group_results):
            month_data = group_results[procurement_current_mi]
            
            # Проверяем есть ли ПРИХОД в текущем месяце (фиксированный или новый)
            arrival_kg = month_data.get("arrive", 0)
            in_transit = month_data.get("in_transit", False)
            containers = month_data.get("containers", 0)
            week_label = month_data.get("wl", "")
            
            # Показываем все приходы (in_transit означает что товар придёт в этом месяце)
            if in_transit and arrival_kg > 0:
                # Определяем автоматическую дату по неделе
                if "пред" in week_label.lower() or "Тиж. 1" in week_label:
                    auto_day = 5
                elif "Тиж. 2" in week_label:
                    auto_day = 10
                elif "Тиж. 3" in week_label:
                    auto_day = 15
                elif "Тиж. 4" in week_label:
                    auto_day = 20
                else:
                    auto_day = 5  # По умолчанию
                
                # Проверяем есть ли зафиксированная дата
                if group_name in st.session_state.arrival_fixed_dates:
                    fixed_date = st.session_state.arrival_fixed_dates[group_name]
                    current_date = fixed_date
                    is_fixed = True
                else:
                    current_date = datetime(current_year, current_month, auto_day)
                    is_fixed = False
                
                orders_current_month.append({
                    "group_name": group_name,
                    "containers": containers,
                    "arrival_kg": arrival_kg,
                    "week_label": week_label,
                    "auto_day": auto_day,
                    "current_date": current_date,
                    "is_fixed": is_fixed,
                    "_arrival_mi": procurement_current_mi  # Месяц прихода для week_arrival
                })
    
    if not orders_current_month:
        st.warning(f"⚠️ Нет заказов с приходом в {current_month_name} {current_year}")
        st.stop()
    
    # Показываем таблицу заказов
    st.markdown(f"### 📦 Заказы {current_month_name} {current_year} ({len(orders_current_month)} шт)")
    
    import pandas as pd
    
    table_data = []
    for order in orders_current_month:
        week_short = ""
        wl = order["week_label"]
        if "пред" in wl.lower():
            week_short = "пред мес"
        elif "Тиж. 1" in wl:
            week_short = "нед 1"
        elif "Тиж. 2" in wl:
            week_short = "нед 2"
        elif "Тиж. 3" in wl:
            week_short = "нед 3"
        elif "Тиж. 4" in wl:
            week_short = "нед 4"
        
        status = "🔒 Зафиксирована" if order["is_fixed"] else "🔄 Автоматическая"
        
        # Формат отображения контейнеров
        if order["containers"] > 0:
            cont_display = f"{order['containers']} конт"
        else:
            cont_display = "баланс"
        
        table_data.append({
            "Группа": order["group_name"],
            "Тип": cont_display,
            "Товар": f"{order['arrival_kg']:,} кг",
            "Неделя": week_short,
            "Дата прихода": order["current_date"].strftime("%d.%m.%Y"),
            "Статус": status
        })
    
    df = pd.DataFrame(table_data)
    st.dataframe(df, use_container_width=True, hide_index=True)
    
    # Форма для фиксации даты
    st.markdown("---")
    st.markdown("### ✏️ Изменить дату прихода")
    
    col1, col2, col3 = st.columns([3, 2, 1])
    
    with col1:
        # Выбор заказа
        order_options = []
        for o in orders_current_month:
            type_str = f"{o['containers']} конт" if o['containers'] > 0 else "баланс"
            order_options.append(f"{o['group_name']} ({type_str}, {o['current_date'].strftime('%d.%m.%Y')})")
        
        selected_option = st.selectbox(
            "Выберите приход:",
            options=order_options,
            key="selected_order"
        )
        selected_idx = order_options.index(selected_option)
        selected_order = orders_current_month[selected_idx]
    
    with col2:
        # Ввод новой даты
        current_value = selected_order["current_date"]
        # Конвертируем datetime в date если нужно
        if isinstance(current_value, datetime):
            current_value = current_value.date()
        
        new_date = st.date_input(
            "Новая дата прихода:",
            value=current_value,
            min_value=datetime(current_year, current_month, 1).date(),
            max_value=datetime(current_year, current_month, 28).date(),
            key="new_arrival_date"
        )
    
    with col3:
        st.write("")  # Отступ
        st.write("")
        if st.button("💾 Зафиксировать", type="primary", use_container_width=True):
            # Сохраняем дату
            fixed_datetime = datetime.combine(new_date, datetime.min.time())
            st.session_state.arrival_fixed_dates[selected_order["group_name"]] = fixed_datetime
            
            # Автоматически вычисляем номер недели из даты
            week_number = (new_date.day - 1) // 7 + 1  # Неделя 1-4
            
            # Находим группу и обновляем week_arrival
            for group in st.session_state.groups:
                if group["name"] == selected_order["group_name"]:
                    # Находим mi для текущего месяца
                    # selected_order содержит данные для текущего месяца
                    arrival_mi = selected_order.get("_arrival_mi", 0)  # mi месяца прихода
                    
                    if "week_arrival" not in group:
                        group["week_arrival"] = {}
                    group["week_arrival"][arrival_mi] = week_number
                    break
            
            # КРИТИЧНО: Пересчитываем симуляцию после изменения
            st.session_state.results = run_all_simulations(st.session_state.groups)
            save_state_to_db()
            st.session_state.need_recalc = False
            
            # Отладка
            st.info(f"🔍 DEBUG: arrival_mi={arrival_mi}, week_number={week_number}, group_name={selected_order['group_name']}")
            
            st.success(f"✅ Дата для '{selected_order['group_name']}' зафиксирована: {new_date.strftime('%d.%m.%Y')} (неделя {week_number})")
            st.rerun()
    
    # Кнопка сброса всех дат
    st.markdown("---")
    
    if st.button("🔄 Сбросить все даты на автоматические", use_container_width=False):
        st.session_state.arrival_fixed_dates = {}
        st.success("✅ Все даты сброшены на автоматические")
        st.rerun()


elif page == "📈 Аналитика по продажам":
    st.title("📈 Аналитика по продажам")
    st.caption("Визуализация динамики продаж и отклонений от плана")
    
    # Проверяем что есть данные
    has_prices = len(st.session_state.get("sales_prices", {})) > 0
    has_plan = len(st.session_state.get("sales_plan_base", {})) > 0
    
    if not has_prices or not has_plan:
        st.warning("⚠️ Сначала импортируйте данные по продажам")
        st.caption("Перейдите в раздел 'Импорт данных по продажам'")
        st.stop()
    
    # Проверяем что симуляция выполнена
    if st.session_state.results is None:
        st.warning("⚠️ Сначала выполните пересчёт закупок")
        st.stop()
    
    # ========================================================================
    # БЛОК 1: МЕТРИКИ ПО ОСТАТКАМ ТЕКУЩЕГО МЕСЯЦА
    # ========================================================================
    st.subheader("📊 Анализ остатков на начало текущего месяца")
    
    # Получаем план продаж
    sales_plan = st.session_state.get("sales_plan_base", {})
    results = st.session_state.results
    
    # Анализируем ТОЛЬКО текущий месяц (mi=0)
    mi = 0
    insufficient_stock_count = 0  # Остаток < план
    critical_stock_count = 0  # Остаток < 50% плана
    total_positions_with_plan = 0  # Всего позиций с планом
    
    # Для списков критичных позиций
    insufficient_list = []
    critical_list = []
    
    for group_idx, group in enumerate(st.session_state.groups):
        # Получаем план продаж для группы (кг/месяц)
        group_sales_plan = sales_plan.get(group_idx, {})
        
        # Для каждой позиции в группе проверяем её остаток
        for item_idx, item in enumerate(group["items"]):
            item_plan = group_sales_plan.get(item_idx, 0)
            
            if item_plan > 0:  # Проверяем только позиции с планом
                total_positions_with_plan += 1
                
                # Берём остаток ПОЗИЦИИ напрямую из balance
                item_stock = item.get("balance", 0)
                
                # Проверяем условия
                if item_stock < item_plan:  # Остаток меньше месячного плана
                    insufficient_stock_count += 1
                    insufficient_list.append({
                        "Группа": group["name"],
                        "Позиция": item["name"],
                        "Остаток": f"{item_stock:,.0f} кг",
                        "План": f"{item_plan:,.0f} кг",
                        "Дефицит": f"{item_plan - item_stock:,.0f} кг",
                        "% обеспеченности": f"{(item_stock / item_plan * 100):.0f}%" if item_plan > 0 else "—"
                    })
                
                if item_stock < item_plan * 0.5:  # Остаток меньше 50% месячного плана
                    critical_stock_count += 1
                    critical_list.append({
                        "Группа": group["name"],
                        "Позиция": item["name"],
                        "Остаток": f"{item_stock:,.0f} кг",
                        "План": f"{item_plan:,.0f} кг",
                        "Дефицит": f"{item_plan - item_stock:,.0f} кг",
                        "% обеспеченности": f"{(item_stock / item_plan * 100):.0f}%" if item_plan > 0 else "—"
                    })
    
    # Показываем метрики для текущего месяца
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric(
            "Позиций с остатком < плана",
            insufficient_stock_count,
            help=f"Количество позиций на начало текущего месяца, остаток которых меньше месячного плана продаж (из {total_positions_with_plan} позиций с планом)"
        )
    
    with col2:
        st.metric(
            "Позиций с остатком < 50% плана",
            critical_stock_count,
            help="Количество позиций на начало текущего месяца, остаток которых меньше половины месячного плана продаж"
        )
    
    with col3:
        # Процент критичных от ВСЕХ позиций с планом
        # Критичные = недостаточные + критичные (все с проблемами)
        total_critical = insufficient_stock_count + critical_stock_count
        if total_positions_with_plan > 0:
            critical_pct = (total_critical / total_positions_with_plan) * 100
        else:
            critical_pct = 0
        
        st.metric(
            "% критичных позиций",
            f"{critical_pct:.0f}%",
            help=f"Процент всех позиций с недостаточным остатком от всех позиций с планом ({total_critical} из {total_positions_with_plan})"
        )
    
    st.divider()
    
    # Таблицы с критичными позициями
    st.subheader("📋 Списки позиций с недостаточным остатком")
    
    tab1, tab2 = st.tabs([
        f"⚠️ Остаток < план ({insufficient_stock_count})",
        f"🔴 Остаток < 50% плана ({critical_stock_count})"
    ])
    
    with tab1:
        if insufficient_list:
            df_insufficient = pd.DataFrame(insufficient_list)
            st.dataframe(df_insufficient, use_container_width=True, hide_index=True, height=400)
        else:
            st.success("✅ Все позиции обеспечены на месяц!")
    
    with tab2:
        if critical_list:
            df_critical = pd.DataFrame(critical_list)
            st.dataframe(df_critical, use_container_width=True, hide_index=True, height=400)
        else:
            st.success("✅ Критичных позиций нет!")


# Футер
st.divider()
st.caption("Система планирования закупок и продаж | v2.0")
